"""
merge_master_logs.py
====================
Merges master_log_price_return_era.xlsx (old, 66 data rows, 49 cols)
with master_log.xlsx (new, 11 data rows, 52 cols) into a single file.

The new file has 3 extra META_COLS the old file lacks:
  - Pricing Model
  - Tx Cost %
  - Tax Drag %

Old rows get blank cells for these three columns.
Old rows that have data in the 10-metric columns keep it.
Old rows that are missing the new metrics (Avg_DD etc.) keep blanks.

Output: results/master_log_merged.xlsx
Run from the project root: python merge_master_logs.py
"""

import os
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

OLD_PATH = os.path.join("results", "master_log_price_return_era.xlsx")
NEW_PATH = os.path.join("results", "master_log.xlsx")
OUT_PATH = os.path.join("results", "master_log_merged.xlsx")

# ── Canonical column structure (matches current export.py exactly) ────────────

META_COLS = [
    "Timestamp", "Label", "Run Mode",
    "Backtest Start", "Backtest End", "OOS Start",
    "Pricing Model", "Tx Cost %", "Tax Drag %",
    "Data Freq", "Tickers",
]

METRIC_COLS = [
    "CAGR (%)", "Max_DD (%)", "Avg_DD (%)", "Max_DD_Dur",
    "Avg_Rec", "Ulcer", "Sharpe", "Sortino", "Calmar", "Fin_Val ($)",
]

STRATEGY_NAMES = ["AW_R", "B&H_AW", "SPY", "60/40"]

STRATEGY_COLOURS = {
    "AW_R":   "1F4E79",
    "B&H_AW": "7B3F00",
    "SPY":    "3D1A1A",
    "60/40":  "1A4731",
}
HEADER_FONT_COLOUR = "FFFFFF"


def flat_columns():
    cols = list(META_COLS)
    for s in STRATEGY_NAMES:
        for m in METRIC_COLS:
            cols.append(f"{s}_{m}")
    cols.append("Results Folder")
    return cols


def read_rows(path, flat_cols):
    """
    Read all data rows from an xlsx file and return as list of dicts
    keyed by flat column names. Handles both old (49-col) and new (52-col)
    formats by matching columns positionally using the header rows.
    """
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    # Build a position -> flat_col_name map from the file's own headers
    # Row 1 = group labels, Row 2 = metric names
    # Meta cols span rows 1+2 (merged), strategy cols have group in row 1
    # and metric name in row 2.
    col_map = {}  # excel_col (1-based) -> flat_col_name

    # Detect meta cols: those where row 2 is empty (merged with row 1)
    r1 = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    r2 = [ws.cell(2, c).value for c in range(1, ws.max_column + 1)]

    current_strategy = None
    for c_idx in range(ws.max_column):
        col = c_idx + 1
        v1 = r1[c_idx]
        v2 = r2[c_idx]

        if v1 in STRATEGY_NAMES:
            current_strategy = v1

        if v2 is None:
            # Meta column or Results Folder (row 1 has the name, row 2 merged)
            if v1 == "Results Folder":
                col_map[col] = "Results Folder"
            elif v1 in META_COLS:
                col_map[col] = v1
            elif v1 in STRATEGY_NAMES:
                # Group header cell itself — skip, handled by row 2 cells
                pass
        else:
            # Metric column — strategy comes from row 1 group header
            if current_strategy and v2 in METRIC_COLS:
                col_map[col] = f"{current_strategy}_{v2}"

    rows = []
    for row_idx in range(3, ws.max_row + 1):
        raw = [ws.cell(row_idx, c).value for c in range(1, ws.max_column + 1)]
        if all(v is None for v in raw):
            continue

        row = {col: "" for col in flat_cols}
        for c_idx, val in enumerate(raw):
            col = c_idx + 1
            if col in col_map and col_map[col] in row:
                if val is not None:
                    row[col_map[col]] = val

        rows.append(row)

    return rows


def write_excel(path, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Master Log"

    col_index = {}
    excel_col = 1

    thin   = Side(style="thin", color="444444")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    meta_style = PatternFill("solid", fgColor="1A1A2E")
    meta_font  = Font(bold=True, color="C9D1D9")
    center     = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Meta columns — merge rows 1+2
    for col_name in META_COLS:
        cell1 = ws.cell(row=1, column=excel_col, value=col_name)
        ws.merge_cells(start_row=1, start_column=excel_col,
                       end_row=2,   end_column=excel_col)
        cell1.fill = meta_style
        cell1.font = meta_font
        cell1.alignment = center
        cell1.border = border
        col_index[col_name] = excel_col
        excel_col += 1

    # Strategy group columns
    for strategy in STRATEGY_NAMES:
        group_start = excel_col
        colour      = STRATEGY_COLOURS.get(strategy, "333333")
        group_fill  = PatternFill("solid", fgColor=colour)
        group_font  = Font(bold=True, color=HEADER_FONT_COLOUR)
        metric_fill = PatternFill("solid", fgColor="161B22")
        metric_font = Font(bold=False, color="C9D1D9")

        for metric in METRIC_COLS:
            flat_name = f"{strategy}_{metric}"
            col_index[flat_name] = excel_col
            cell2 = ws.cell(row=2, column=excel_col, value=metric)
            cell2.fill = metric_fill
            cell2.font = metric_font
            cell2.alignment = center
            cell2.border = border
            excel_col += 1

        group_end = excel_col - 1
        cell1 = ws.cell(row=1, column=group_start, value=strategy)
        ws.merge_cells(start_row=1, start_column=group_start,
                       end_row=1,   end_column=group_end)
        cell1.fill = group_fill
        cell1.font = group_font
        cell1.alignment = center
        cell1.border = border

    # Results Folder
    rf_col = excel_col
    cell1  = ws.cell(row=1, column=rf_col, value="Results Folder")
    ws.merge_cells(start_row=1, start_column=rf_col,
                   end_row=2,   end_column=rf_col)
    cell1.fill = meta_style
    cell1.font = meta_font
    cell1.alignment = center
    cell1.border = border
    col_index["Results Folder"] = rf_col

    # Data rows
    data_fill_even = PatternFill("solid", fgColor="FFFFFF")
    data_fill_odd  = PatternFill("solid", fgColor="F2F2F2")
    data_font      = Font(color="000000")
    num_align      = Alignment(horizontal="right")
    left_align     = Alignment(horizontal="left")

    for row_idx, row_data in enumerate(rows):
        excel_row = row_idx + 3
        fill = data_fill_even if row_idx % 2 == 0 else data_fill_odd

        for flat_col, ec in col_index.items():
            value = row_data.get(flat_col, "")
            if isinstance(value, float) and value != value:
                value = ""
            cell = ws.cell(row=excel_row, column=ec, value=value)
            cell.fill = fill
            cell.font = data_font
            cell.border = border
            if isinstance(value, (int, float)) and value != "":
                cell.alignment = num_align
                if isinstance(value, float):
                    cell.number_format = "0.000"
            else:
                cell.alignment = left_align

    # Column widths
    width_map = {
        "Timestamp":      19,
        "Label":          30,
        "Run Mode":       14,
        "Backtest Start": 13,
        "Backtest End":   13,
        "OOS Start":      13,
        "Pricing Model":  13,
        "Tx Cost %":      10,
        "Tax Drag %":     10,
        "Data Freq":      10,
        "Tickers":        55,
        "Results Folder": 60,
    }
    for col_name, width in width_map.items():
        if col_name in col_index:
            ws.column_dimensions[get_column_letter(col_index[col_name])].width = width

    metric_widths = {
        "CAGR (%)":   11, "Max_DD (%)": 13, "Avg_DD (%)": 13,
        "Max_DD_Dur": 13, "Avg_Rec":    10, "Ulcer":       9,
        "Sharpe":      9, "Sortino":    10, "Calmar":      9,
        "Fin_Val ($)": 14,
    }
    for s in STRATEGY_NAMES:
        for m in METRIC_COLS:
            flat = f"{s}_{m}"
            if flat in col_index:
                ws.column_dimensions[
                    get_column_letter(col_index[flat])
                ].width = metric_widths.get(m, 11)

    freeze_col = get_column_letter(len(META_COLS) + 1)
    ws.freeze_panes = f"{freeze_col}3"

    wb.save(path)
    print(f"Written -> {path}  ({len(rows)} data rows, {len(col_index)} columns)")


# ── Main ─────────────────────────────────────────────────────────────────────

for p in [OLD_PATH, NEW_PATH]:
    if not os.path.exists(p):
        print(f"ERROR: {p} not found. Run from the project root.")
        exit(1)

flat_cols = flat_columns()

print(f"Reading old log: {OLD_PATH}")
old_rows = read_rows(OLD_PATH, flat_cols)
print(f"  {len(old_rows)} rows read")

print(f"Reading new log: {NEW_PATH}")
new_rows = read_rows(NEW_PATH, flat_cols)
print(f"  {len(new_rows)} rows read")

# Old rows come first (chronologically earlier), new rows appended after
all_rows = old_rows + new_rows
print(f"\nTotal rows to write: {len(all_rows)}")

write_excel(OUT_PATH, all_rows)

print(f"\nMerge complete.")
print(f"  Output -> {OUT_PATH}")
print(f"\nVerify in Excel:")
print(f"  - First {len(old_rows)} rows: old runs, blank Pricing Model / Tx Cost % / Tax Drag %")
print(f"  - Last {len(new_rows)} rows: new runs, all columns populated")
print(f"  - Results Folder always in the last column")
print(f"  - No data from either file lost")
