"""
research/compare_jepq.py
========================
Head-to-head comparison of the All-Weather strategy vs JEPQ buy-and-hold
over the full JEPQ price history (inception 2022-05-03 → today).

JEPQ context
------------
JEPQ (JPMorgan Nasdaq Equity Premium Income ETF) sells covered calls on
the Nasdaq-100 to generate monthly income (~8–11% p.a. yield).  It
captures most of the QQQ upside in bull markets but lags badly on large
moves because the covered calls cap gains.  It is included as a benchmark
to show how a yield-focused equity strategy compares against a risk-parity
approach over the same window.

Outputs
-------
  - Side-by-side performance table (stdout)
  - results/jepq_comparison_growth.png
  - results/jepq_comparison_<date>.xlsx

Usage
-----
  conda run -n allweather python -m research.compare_jepq

Notes
-----
* JEPQ pays large monthly distributions.  yfinance `auto_adjust=True`
  adjusts for dividends so the price return here includes reinvested income.
  Total-return is the correct comparison for an income-focused ETF.
* Inception: 2022-05-03 (first trading day).
* Expense ratio: 0.35% p.a.
"""

from __future__ import annotations

import os
import sys
import warnings
from datetime import date

import matplotlib.dates as mdates
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import numpy as np
import pandas as pd
import yfinance as yf

warnings.filterwarnings("ignore", category=FutureWarning)
warnings.filterwarnings("ignore", message=".*auto_adjust.*")

_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_PROJECT_ROOT = os.path.dirname(_SCRIPT_DIR)
sys.path.insert(0, _PROJECT_ROOT)

from engine.stats import compute_cagr, compute_max_drawdown, compute_sharpe, compute_calmar
from research.compare_allw import (
    _alloc_from_json,
    _strip_tz,
    apply_annual_fee,
    build_daily_series,
)

_RESULTS_DIR = os.path.join(_SCRIPT_DIR, "results")
os.makedirs(_RESULTS_DIR, exist_ok=True)

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

JEPQ_INCEPTION   = "2022-05-03"
DATE_END         = date.today().strftime("%Y-%m-%d")

STRATEGY_FEE     = 0.0012   # ~0.12% p.a. weighted-avg ETF expenses
JEPQ_FEE         = 0.0035   # 0.35% p.a.

DARK_BG  = "#1e1e2e"
DARK_FG  = "#cdd6f4"
GRID_CLR = "#313244"

# ---------------------------------------------------------------------------
# Data helpers
# ---------------------------------------------------------------------------

def _fetch_prices(start: str, end: str) -> pd.DataFrame:
    tickers = ["SPY", "TLT", "QQQ", "JEPQ"]
    # Also fetch strategy tickers
    try:
        alloc = _alloc_from_json("6asset_tip_gsg_rpavg", use_live=True)
        tickers = sorted(set(tickers) | set(alloc.keys()))
    except Exception:
        alloc = {}

    print(f"Fetching {' '.join(tickers)}  [{start} → {end}]")
    frames = {}
    for tkr in tickers:
        try:
            hist = yf.Ticker(tkr).history(start=start, end=end, auto_adjust=True)
            if hist.empty:
                continue
            hist.index = _strip_tz(hist.index)
            frames[tkr] = hist["Close"].rename(tkr)
        except Exception as exc:
            print(f"  WARNING: could not fetch {tkr}: {exc}")

    if "JEPQ" not in frames:
        raise SystemExit(
            "JEPQ price data is unavailable.  "
            "Check your internet connection or try again later."
        )

    prices = pd.DataFrame(frames).dropna(how="all").ffill()
    print(f"  {len(prices)} trading days  "
          f"({prices.index[0].date()} → {prices.index[-1].date()})\n")
    return prices, alloc


# ---------------------------------------------------------------------------
# Stats helper
# ---------------------------------------------------------------------------

def _stats(series: pd.Series, label: str) -> dict:
    daily_ret = series.pct_change().dropna()
    total_return = (series.iloc[-1] / series.iloc[0] - 1) * 100
    n_days = len(series)
    n_years = n_days / 252
    cagr = compute_cagr(series.iloc[0], series.iloc[-1], n_years) * 100
    mdd = compute_max_drawdown(series) * 100
    sharpe = compute_sharpe(daily_ret) if len(daily_ret) > 20 else float("nan")
    calmar = compute_calmar(cagr / 100, mdd / 100)
    return {
        "Strategy": label,
        "Period": f"{series.index[0].date()} → {series.index[-1].date()}",
        "Days": n_days,
        "Total Ret%": round(total_return, 2),
        "CAGR%": round(cagr, 2),
        "Max DD%": round(mdd, 2),
        "Sharpe": round(sharpe, 3),
        "Calmar": round(calmar, 3),
    }


# ---------------------------------------------------------------------------
# Table output
# ---------------------------------------------------------------------------

def _print_table(rows: list[dict]) -> None:
    df = pd.DataFrame([r for r in rows if r is not None])
    col_w = {c: max(len(c), df[c].astype(str).str.len().max()) for c in df.columns}
    header = "  ".join(c.ljust(col_w[c]) for c in df.columns)
    sep = "  ".join("-" * col_w[c] for c in df.columns)
    print(header)
    print(sep)
    for _, row in df.iterrows():
        print("  ".join(str(row[c]).ljust(col_w[c]) for c in df.columns))
    print()


# ---------------------------------------------------------------------------
# Chart
# ---------------------------------------------------------------------------

def _chart(series_map: dict[str, pd.Series], title: str) -> None:
    fig, ax = plt.subplots(figsize=(12, 6.5))
    fig.patch.set_facecolor(DARK_BG)
    ax.set_facecolor(DARK_BG)
    ax.tick_params(colors=DARK_FG)
    ax.yaxis.label.set_color(DARK_FG)
    ax.xaxis.label.set_color(DARK_FG)
    ax.title.set_color(DARK_FG)
    for spine in ax.spines.values():
        spine.set_edgecolor(GRID_CLR)
    ax.grid(color=GRID_CLR, linewidth=0.5, linestyle="--", alpha=0.6)

    colours = {
        "AW (6-asset RP)": "#89b4fa",
        "JEPQ (total ret)": "#c678dd",
        "QQQ":              "#f0b429",
        "SPY":              "#f78166",
        "60/40":            "#3fb950",
    }
    styles = {
        "AW (6-asset RP)":  ("-",  2.2),
        "JEPQ (total ret)": ("--", 2.2),
        "QQQ":              ("-.", 1.5),
        "SPY":              (":",  1.5),
        "60/40":            ("-.", 1.5),
    }

    # Align all series to the same start date (common index)
    common_start = max(s.index[0] for s in series_map.values())
    for label, s in series_map.items():
        s = s.loc[s.index >= common_start]
        s = s / s.iloc[0] * 100.0
        ls, lw = styles.get(label, ("-", 1.5))
        ax.plot(s.index, s, label=label,
                color=colours.get(label, "#aaaaaa"),
                linewidth=lw, linestyle=ls)

    ax.set_title(title, color=DARK_FG, fontsize=13, pad=12)
    ax.set_ylabel("Growth of $100", color=DARK_FG)
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f"${x:,.0f}"))
    ax.xaxis.set_major_formatter(mdates.DateFormatter("%Y-%m"))
    plt.xticks(rotation=30)
    ax.legend(facecolor="#313244", labelcolor=DARK_FG, fontsize=9, framealpha=0.8)
    plt.tight_layout()

    fname = f"{date.today().strftime('%Y-%m-%d')}_jepq_comparison_growth"
    for ext in (".png", ".svg"):
        path = os.path.join(_RESULTS_DIR, fname + ext)
        plt.savefig(path, dpi=100, bbox_inches="tight", facecolor=fig.get_facecolor())
        print(f"Saved: {path}")
    plt.close(fig)


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

def _save_excel(rows: list[dict]) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        print("openpyxl not installed — skipping Excel export.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "JEPQ Comparison"
    df = pd.DataFrame([r for r in rows if r is not None])
    ws.append(list(df.columns))
    for _, row in df.iterrows():
        ws.append(list(row))

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="1a1a2e")
        cell.font = Font(bold=True, color="89b4fa")

    for col in ws.columns:
        width = max(len(str(cell.value or "")) for cell in col) + 2
        ws.column_dimensions[col[0].column_letter].width = width

    fname = os.path.join(_RESULTS_DIR,
                         f"{date.today().strftime('%Y-%m-%d')}_jepq_comparison.xlsx")
    wb.save(fname)
    print(f"Excel saved: {fname}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    prices, alloc = _fetch_prices(JEPQ_INCEPTION, DATE_END)

    # All-Weather rebalanced
    aw_gross = build_daily_series(prices, alloc, start_value=100.0, rebalance=True)
    aw_fa    = apply_annual_fee(aw_gross, STRATEGY_FEE)

    # JEPQ buy-and-hold (total-return, includes dividends via auto_adjust)
    jepq_gross = prices["JEPQ"] / prices["JEPQ"].iloc[0] * 100.0
    jepq_fa    = apply_annual_fee(jepq_gross, JEPQ_FEE)

    # SPY buy-and-hold
    spy = prices["SPY"] / prices["SPY"].iloc[0] * 100.0

    # QQQ buy-and-hold (context: JEPQ tracks Nasdaq-100)
    qqq = None
    if "QQQ" in prices.columns:
        qqq = prices["QQQ"] / prices["QQQ"].iloc[0] * 100.0

    # 60/40
    s_6040 = None
    if "TLT" in prices.columns:
        spy_sh = 60 / float(prices["SPY"].iloc[0])
        tlt_sh = 40 / float(prices["TLT"].iloc[0])
        s_6040 = (spy_sh * prices["SPY"] + tlt_sh * prices["TLT"])
        s_6040 = s_6040 / s_6040.iloc[0] * 100.0

    # ── Print table ───────────────────────────────────────────────────────────
    rows: list[dict | None] = [
        _stats(aw_gross,  "AW 6-asset RP  (monthly rebal, gross)"),
        _stats(aw_fa,     "  → fee-adj 0.12% p.a."),
        None,
        _stats(jepq_gross, "JEPQ  (JPM Nasdaq covered-call, gross)"),
        _stats(jepq_fa,    "  → fee-adj 0.35% p.a."),
        None,
        _stats(spy,        "SPY  (S&P 500 buy-and-hold)"),
    ]
    if qqq is not None:
        rows.append(_stats(qqq, "QQQ  (Nasdaq-100 buy-and-hold)"))
    if s_6040 is not None:
        rows.append(_stats(s_6040, "60/40  (SPY/TLT rebalanced)"))

    print("\n" + "=" * 80)
    print("  AW vs JEPQ  —  full JEPQ history since inception 2022-05-03")
    print("=" * 80)
    _print_table(rows)

    # ── Chart ─────────────────────────────────────────────────────────────────
    series_map = {
        "AW (6-asset RP)":  aw_fa,
        "JEPQ (total ret)": jepq_fa,
        "SPY":              spy,
    }
    if qqq is not None:
        series_map["QQQ"] = qqq
    if s_6040 is not None:
        series_map["60/40"] = s_6040

    _chart(series_map, title=(
        "All-Weather 6-Asset RP vs JEPQ  —  "
        f"since JEPQ inception {JEPQ_INCEPTION}"
    ))

    # ── Excel ─────────────────────────────────────────────────────────────────
    _save_excel(rows)

    print("\nDone.")


if __name__ == "__main__":
    main()
