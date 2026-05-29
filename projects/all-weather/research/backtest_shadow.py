"""
research/backtest_shadow.py
===========================
Backtest shadow reconciliation — compare actual live returns recorded in
logs/performance_tracking_*.csv against the engine-simulated returns for
the same calendar period.

Why
---
The live rebalancer uses real fills, bid-ask spread, and actual dividend
timing.  The engine simulation assumes perfect monthly rebalances at
month-end closing prices.  This script quantifies the gap.

Outputs
-------
  - Side-by-side table: Actual vs Simulated monthly returns
  - Cumulative deviation chart
  - results/backtest_shadow_<date>.xlsx

Usage
-----
  conda run -n allweather python -m research.backtest_shadow
  conda run -n allweather python -m research.backtest_shadow \
      --csv logs/performance_tracking_alpaca_paper_default.csv \
      --strategy-id 6asset_tip_gsg_rpavg
"""

from __future__ import annotations

import argparse
import glob
import os
import sys
from datetime import date, datetime
from typing import Any

import numpy as np
import pandas as pd
import yfinance as yf

_SCRIPT_DIR    = os.path.dirname(os.path.abspath(__file__))
_PROJECT_ROOT  = os.path.dirname(_SCRIPT_DIR)
sys.path.insert(0, _PROJECT_ROOT)

from engine.stats import compute_cagr, compute_max_drawdown, compute_sharpe, compute_calmar
from research.compare_allw import _alloc_from_json, _strip_tz

_RESULTS_DIR = os.path.join(_SCRIPT_DIR, "results")
_LOGS_DIR    = os.path.join(_PROJECT_ROOT, "logs")
os.makedirs(_RESULTS_DIR, exist_ok=True)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _find_performance_csv(logs_dir: str) -> list[str]:
    """Return all performance_tracking_*.csv files found in logs/."""
    return sorted(glob.glob(os.path.join(logs_dir, "performance_tracking_*.csv")))


def _load_actual(csv_path: str) -> pd.DataFrame:
    """Load the live performance CSV into a clean DataFrame indexed by date."""
    df = pd.read_csv(csv_path)
    df["Date"] = pd.to_datetime(df["Date"])
    df = df.set_index("Date").sort_index()
    # Normalise equity column: may be "$100,000.00" or plain float
    if "Portfolio_Equity" in df.columns:
        df["Portfolio_Equity"] = (
            df["Portfolio_Equity"]
            .astype(str)
            .str.replace(r"[\$,]", "", regex=True)
            .astype(float)
        )
    return df


def _simulate_monthly_returns(
    allocation: dict[str, float],
    start: str,
    end: str,
) -> pd.Series:
    """Run the engine backtest over the given date range and return monthly returns."""
    from engine.backtest import run_backtest
    from engine.calendar import pandas_resample_frequency
    from engine import config

    tickers = list(allocation.keys())
    spy_ticker = "SPY"

    all_tickers = sorted(set(tickers + [spy_ticker]))
    print(f"  Fetching simulation prices: {' '.join(all_tickers)}  [{start} → {end}]")

    raw = yf.download(all_tickers, start=start, end=end,
                      auto_adjust=True, progress=False)
    closes = raw["Close"] if "Close" in raw else raw
    if isinstance(closes, pd.Series):
        closes = closes.to_frame()
    closes.index = _strip_tz(closes.index)
    closes = closes.dropna(how="all").ffill()

    prices = closes[tickers].dropna()
    bench  = closes[spy_ticker].dropna()

    df = run_backtest(
        prices=prices,
        benchmark_prices=bench,
        allocation=allocation,
        portfolio_value=100.0,
    )
    # Extract monthly portfolio returns
    monthly_ret = df["All Weather Value Monthly Ret (%)"].dropna()
    return monthly_ret


# ---------------------------------------------------------------------------
# Main comparison
# ---------------------------------------------------------------------------

def compare(csv_path: str, strategy_id: str) -> None:
    print(f"\nBacktest shadow reconciliation")
    print(f"  Live log:    {csv_path}")
    print(f"  Strategy:    {strategy_id}")
    print()

    # 1. Load actual returns
    actual_df = _load_actual(csv_path)
    if "Portfolio_Return%" not in actual_df.columns:
        raise SystemExit(
            f"Column 'Portfolio_Return%' not found in {csv_path}. "
            "The rebalancer must have been run at least twice to record returns."
        )

    actual_ret = actual_df["Portfolio_Return%"].dropna()
    if actual_ret.empty:
        raise SystemExit("No non-null return rows found in the performance CSV.")

    start_str = (actual_ret.index[0] - pd.DateOffset(months=2)).strftime("%Y-%m-%d")
    end_str   = (actual_ret.index[-1] + pd.DateOffset(days=5)).strftime("%Y-%m-%d")

    # 2. Load allocation
    try:
        allocation = _alloc_from_json(strategy_id, use_live=True)
    except Exception as exc:
        raise SystemExit(f"Could not load allocation for {strategy_id}: {exc}")

    # 3. Simulate monthly returns over same window
    sim_ret = _simulate_monthly_returns(allocation, start_str, end_str)

    # 4. Align on month-end dates (match by year-month)
    actual_df_reind = actual_ret.to_frame("Actual_%")
    actual_df_reind["YM"] = actual_df_reind.index.to_period("M")

    sim_df = sim_ret.to_frame("Sim_%")
    sim_df.index = pd.to_datetime(sim_df.index)
    sim_df["YM"] = sim_df.index.to_period("M")

    merged = pd.merge(
        actual_df_reind.reset_index(),
        sim_df.reset_index()[["YM", "Sim_%"]],
        on="YM",
        how="inner",
    ).set_index("Date").drop(columns="YM")

    if merged.empty:
        print("WARNING: No overlapping month-year periods between actual and simulated data.")
        return

    merged["Delta_%"] = merged["Actual_%"] - merged["Sim_%"]

    # 5. Print table
    print(merged.round(3).to_string())
    print()

    # 6. Summary stats
    mae   = merged["Delta_%"].abs().mean()
    rmse  = (merged["Delta_%"] ** 2).mean() ** 0.5
    bias  = merged["Delta_%"].mean()
    print(f"Periods compared : {len(merged)}")
    print(f"Mean abs error   : {mae:.3f}%")
    print(f"RMSE             : {rmse:.3f}%")
    print(f"Bias (live-sim)  : {bias:+.3f}%  "
          f"({'live outperforms sim' if bias > 0 else 'sim outperforms live'})")
    print()

    # 7. Cumulative deviation chart
    _chart_cumulative_deviation(merged, strategy_id, csv_path)

    # 8. Excel export
    _save_excel(merged, strategy_id)


def _chart_cumulative_deviation(
    merged: pd.DataFrame,
    strategy_id: str,
    csv_path: str,
) -> None:
    try:
        import matplotlib.pyplot as plt
        import matplotlib.dates as mdates
    except ImportError:
        print("matplotlib not available — skipping chart.")
        return

    DARK_BG = "#1e1e2e"
    DARK_FG = "#cdd6f4"
    GRID    = "#313244"

    cum_actual = (1 + merged["Actual_%"] / 100).cumprod()
    cum_sim    = (1 + merged["Sim_%"]    / 100).cumprod()
    cum_delta  = cum_actual - cum_sim

    fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(12, 8), sharex=True)
    fig.patch.set_facecolor(DARK_BG)

    for ax in (ax1, ax2):
        ax.set_facecolor(DARK_BG)
        ax.tick_params(colors=DARK_FG, labelsize=9)
        ax.grid(color=GRID, linewidth=0.4, linestyle="--", alpha=0.5)
        for spine in ax.spines.values():
            spine.set_edgecolor(GRID)

    ax1.plot(merged.index, cum_actual, label="Actual (live)", color="#89b4fa", linewidth=2)
    ax1.plot(merged.index, cum_sim,    label="Simulated",     color="#a6e3a1", linewidth=2,
             linestyle="--")
    ax1.set_ylabel("Cumulative growth (×1)", color=DARK_FG)
    ax1.set_title(f"Backtest Shadow — {strategy_id}", color=DARK_FG, fontsize=12)
    ax1.legend(facecolor="#313244", labelcolor=DARK_FG, fontsize=9)

    ax2.bar(merged.index, cum_delta * 100,
            color=["#89b4fa" if v >= 0 else "#f38ba8" for v in cum_delta],
            width=15, alpha=0.8)
    ax2.axhline(0, color=DARK_FG, linewidth=0.8)
    ax2.set_ylabel("Cumulative deviation (%)", color=DARK_FG)
    ax2.xaxis.set_major_formatter(mdates.DateFormatter("%Y-%m"))
    plt.xticks(rotation=30)

    plt.tight_layout()
    today = date.today().strftime("%Y-%m-%d")
    for ext in (".png", ".svg"):
        path = os.path.join(_RESULTS_DIR, f"{today}_backtest_shadow{ext}")
        plt.savefig(path, dpi=100, bbox_inches="tight", facecolor=fig.get_facecolor())
        print(f"Saved: {path}")
    plt.close(fig)


def _save_excel(merged: pd.DataFrame, strategy_id: str) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Shadow Reconciliation"
    ws.append(["Date"] + list(merged.columns))
    for idx, row in merged.iterrows():
        ws.append([str(idx.date())] + [round(v, 4) for v in row])

    for cell in ws[1]:
        cell.font = Font(bold=True, color="89b4fa")
        cell.fill = PatternFill("solid", fgColor="1a1a2e")

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18

    today = date.today().strftime("%Y-%m-%d")
    safe  = strategy_id.replace("/", "_")
    fname = os.path.join(_RESULTS_DIR, f"{today}_backtest_shadow_{safe}.xlsx")
    wb.save(fname)
    print(f"Excel: {fname}")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description="Backtest shadow: compare live returns to engine simulation."
    )
    p.add_argument(
        "--csv", default=None,
        help=(
            "Path to performance_tracking_*.csv. "
            "Omit to auto-detect from logs/ (first match used)."
        ),
    )
    p.add_argument(
        "--strategy-id", default="6asset_tip_gsg_rpavg",
        help="Strategy id to load allocation from strategies.json.",
    )
    return p.parse_args()


def main() -> None:
    args = parse_args()

    csv_path = args.csv
    if csv_path is None:
        candidates = _find_performance_csv(_LOGS_DIR)
        if not candidates:
            raise SystemExit(
                f"No performance_tracking_*.csv found in {_LOGS_DIR}. "
                "Run the rebalancer with --execute at least twice first."
            )
        csv_path = candidates[0]
        print(f"Auto-detected CSV: {csv_path}")

    compare(csv_path, args.strategy_id)
    print("Done.")


if __name__ == "__main__":
    main()
