"""
migrate_log.py
==============
One-off script to migrate the existing master_log.csv into the new
master_log.xlsx format with grouped strategy headers.

Run once from the project root:
    python migrate_log.py

What it does:
  1. Reads results/master_log.csv
  2. Renames all columns to the new short names
  3. Writes results/master_log.xlsx with grouped headers
  4. Keeps the original CSV as results/master_log_backup.csv

After running this, delete or archive master_log.csv -- future runs
will write to master_log.xlsx only.
"""

import os
import shutil
import pandas as pd

# ---------------------------------------------------------------------------
# Column rename map: old CSV name -> new flat name used by export.py
# ---------------------------------------------------------------------------
COLUMN_RENAMES = {
    "Data Frequency": "Data Freq",

    # All Weather Rebalanced
    "All_Weather_(Rebalanced)_CAGR (%)":        "AW_R_CAGR (%)",
    "All_Weather_(Rebalanced)_Max_DD (%)":       "AW_R_Max_DD (%)",
    "All_Weather_(Rebalanced)_Sharpe":           "AW_R_Sharpe",
    "All_Weather_(Rebalanced)_Calmar":           "AW_R_Calmar",
    "All_Weather_(Rebalanced)_Final_Value ($)":  "AW_R_Fin_Val ($)",

    # Buy & Hold All Weather
    "Buy_and_Hold_All_Weather_CAGR (%)":         "B&H_AW_CAGR (%)",
    "Buy_and_Hold_All_Weather_Max_DD (%)":        "B&H_AW_Max_DD (%)",
    "Buy_and_Hold_All_Weather_Sharpe":            "B&H_AW_Sharpe",
    "Buy_and_Hold_All_Weather_Calmar":            "B&H_AW_Calmar",
    "Buy_and_Hold_All_Weather_Final_Value ($)":   "B&H_AW_Fin_Val ($)",

    # S&P 500
    "SandP_500_Buy_and_Hold_CAGR (%)":           "SPY_CAGR (%)",
    "SandP_500_Buy_and_Hold_Max_DD (%)":          "SPY_Max_DD (%)",
    "SandP_500_Buy_and_Hold_Sharpe":              "SPY_Sharpe",
    "SandP_500_Buy_and_Hold_Calmar":              "SPY_Calmar",
    "SandP_500_Buy_and_Hold_Final_Value ($)":     "SPY_Fin_Val ($)",
}

META_COLS      = ["Timestamp", "Label", "Backtest Start", "Backtest End",
                  "Data Freq", "Tickers"]
STRATEGY_NAMES = ["AW_R", "B&H_AW", "SPY"]
METRIC_COLS    = ["CAGR (%)", "Max_DD (%)", "Sharpe", "Calmar", "Fin_Val ($)"]

STRATEGY_COLOURS = {
    "AW_R":   "1F4E79",
    "B&H_AW": "7B3F00",
    "SPY":    "3D1A1A",
}
HEADER_FONT_COLOUR = "FFFFFF"


def all_flat_columns():
    cols = list(META_COLS)
    for strategy in STRATEGY_NAMES:
        for metric in METRIC_COLS:
            cols.append(f"{strategy}_{metric}")
    cols.append("Results Folder")
    return cols


def write_excel_log(log_path: str, rows: list):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Master Log"

    col_index = {}
    excel_col = 1

    thin   = Side(style="thin", color="444444")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    meta_style = PatternFill("solid", fgColor="1A1A2E")
    meta_font  = Font(bold=True, color="C9D1D9")
    center     = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_name in META_COLS:
        cell1 = ws.cell(row=1, column=excel_col, value=col_name)
        ws.merge_cells(start_row=1, start_column=excel_col,
                       end_row=2,   end_column=excel_col)
        cell1.fill      = meta_style
        cell1.font      = meta_font
        cell1.alignment = center
        cell1.border    = border
        col_index[col_name] = excel_col
        excel_col += 1

    for strategy in STRATEGY_NAMES:
        group_start  = excel_col
        colour       = STRATEGY_COLOURS.get(strategy, "333333")
        group_fill   = PatternFill("solid", fgColor=colour)
        group_font   = Font(bold=True, color=HEADER_FONT_COLOUR)
        metric_fill  = PatternFill("solid", fgColor="161B22")
        metric_font  = Font(bold=False, color="C9D1D9")

        for metric in METRIC_COLS:
            flat_name = f"{strategy}_{metric}"
            col_index[flat_name] = excel_col
            cell2 = ws.cell(row=2, column=excel_col, value=metric)
            cell2.fill      = metric_fill
            cell2.font      = metric_font
            cell2.alignment = center
            cell2.border    = border
            excel_col += 1

        group_end = excel_col - 1
        cell1 = ws.cell(row=1, column=group_start, value=strategy)
        ws.merge_cells(start_row=1, start_column=group_start,
                       end_row=1,   end_column=group_end)
        cell1.fill      = group_fill
        cell1.font      = group_font
        cell1.alignment = center
        cell1.border    = border

    rf_col = excel_col
    cell1  = ws.cell(row=1, column=rf_col, value="Results Folder")
    ws.merge_cells(start_row=1, start_column=rf_col,
                   end_row=2,   end_column=rf_col)
    cell1.fill      = meta_style
    cell1.font      = meta_font
    cell1.alignment = center
    cell1.border    = border
    col_index["Results Folder"] = rf_col

    data_fill_even = PatternFill("solid", fgColor="FFFFFF")
    data_fill_odd  = PatternFill("solid", fgColor="F2F2F2")
    data_font      = Font(color="000000")
    num_align      = Alignment(horizontal="right")
    left_align     = Alignment(horizontal="left")

    for row_idx, row_data in enumerate(rows):
        excel_row = row_idx + 3
        fill      = data_fill_even if row_idx % 2 == 0 else data_fill_odd
        for flat_col, ec in col_index.items():
            value = row_data.get(flat_col, "")
            if isinstance(value, float) and pd.isna(value):
                value = ""
            cell  = ws.cell(row=excel_row, column=ec, value=value)
            cell.fill   = fill
            cell.font   = data_font
            cell.border = border
            if isinstance(value, (int, float)) and value != "":
                cell.alignment = num_align
                if isinstance(value, float):
                    cell.number_format = "0.000"
            else:
                cell.alignment = left_align

    ws.column_dimensions[get_column_letter(col_index["Timestamp"])].width      = 16
    ws.column_dimensions[get_column_letter(col_index["Label"])].width          = 30
    ws.column_dimensions[get_column_letter(col_index["Backtest Start"])].width = 13
    ws.column_dimensions[get_column_letter(col_index["Backtest End"])].width   = 13
    ws.column_dimensions[get_column_letter(col_index["Data Freq"])].width      = 10
    ws.column_dimensions[get_column_letter(col_index["Tickers"])].width        = 45
    ws.column_dimensions[get_column_letter(col_index["Results Folder"])].width = 55
    for strategy in STRATEGY_NAMES:
        for metric in METRIC_COLS:
            flat = f"{strategy}_{metric}"
            ws.column_dimensions[get_column_letter(col_index[flat])].width = 11

    freeze_col = get_column_letter(len(META_COLS) + 1)
    ws.freeze_panes = f"{freeze_col}3"

    wb.save(log_path)
    print(f"Excel log written -> {log_path}")


# ---------------------------------------------------------------------------
# Main migration
# ---------------------------------------------------------------------------

CSV_PATH    = os.path.join("results", "master_log.csv")
BACKUP_PATH = os.path.join("results", "master_log_backup.csv")
XLSX_PATH   = os.path.join("results", "master_log.xlsx")

if not os.path.exists(CSV_PATH):
    print(f"ERROR: {CSV_PATH} not found. Run from the project root directory.")
    exit(1)

shutil.copy(CSV_PATH, BACKUP_PATH)
print(f"Backup saved         -> {BACKUP_PATH}")

df = pd.read_csv(CSV_PATH)
print(f"\nFound {len(df)} existing rows.")
print(f"Columns before rename: {list(df.columns)}")

df = df.rename(columns=COLUMN_RENAMES)
print(f"Columns after rename:  {list(df.columns)}")

flat_cols    = all_flat_columns()
missing_cols = [c for c in flat_cols if c not in df.columns]
if missing_cols:
    print(f"\nWARNING: These expected columns were not found and will be blank: {missing_cols}")
    for c in missing_cols:
        df[c] = ""

df   = df.reindex(columns=flat_cols)
rows = df.to_dict(orient="records")
write_excel_log(XLSX_PATH, rows)

print(f"\nMigration complete.")
print(f"  Old CSV backed up -> {BACKUP_PATH}")
print(f"  New Excel file    -> {XLSX_PATH}")
print(f"\nYou can now delete master_log.csv if everything looks correct.")