"""
curate_master_log.py
====================
Produces master_log_curated.xlsx with two tabs:

  Tab 1 "Curated"  -- one row per experiment per step, deduplicated,
                      old-era rows excluded, sorted by OOS Calmar desc.
                      This is the decision-making view.

  Tab 2 "Archive"  -- every row from master_log.xlsx untouched.
                      Full reproducibility record.

Curation rules (professional standard):
  1. Old-era rows (pre-Phase5, rows without the exp_ prefix and without
     the new 10-metric structure) go to Archive only.
  2. For duplicate steps within the same experiment (e.g. 5asset_dalio
     was run 3 times), keep only the LAST run (most recent timestamp).
  3. Spot-check experiments keep only the single spot_full row.
  4. The following intermediate steps are excluded from Curated:
       s1_backtest  -- IS baseline, superseded by s2_optimise
       s2_optimise  -- IS optimise, useful reference but not the headline
     Unless: no s4_oos row exists (experiment was IS-only), in which case
     keep s1_backtest as the only record.
  5. Keep all of: s2_optimise, s3_walkforward (not in log but in csvs),
     s4_oos, s5_full, spot_full.
  6. Sort Curated tab: first by experiment group (A-E), then by OOS Calmar
     descending within group. Unvalidated experiments (no OOS) go last.

Run from the project root:
    python3 curate_master_log.py
"""

import os
from datetime import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (Alignment, Font, PatternFill, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter

SRC  = os.path.join("results", "master_log.xlsx")
OUT  = os.path.join("results", "master_log_curated.xlsx")

# Experiments that are fully validated (IS + OOS complete)
VALIDATED = {
    "6asset_tip_gsg",
    "6asset_tip_gsg_split2018",
    "6asset_tip_gsg_split2022",
    "8asset_manual_v1",          # from direct oos_evaluate row
    "5asset_dalio_original",
    "5asset_tip_previous_best",
    "7asset_tip_gsg_vnq",
    "7asset_no_iwd",
    "7asset_tip_gsg_dual",
    "8asset_djp_replaces_gsg",
    "8asset_tip_replaces_shy",
    "8asset_efa_replaces_iwd",
    "6asset_duration_ladder",
    "6asset_intl_equity",
    "7asset_no_gsg",
    "7asset_vnq_reits",
    "8asset_agg_replaces_qqq",
    "8asset_lqd_replaces_shy",
}

# Steps to keep in curated tab
KEEP_STEPS = {"s2_optimise", "s4_oos", "s5_full", "spot_full"}

# Group ordering for sorting
GROUP_ORDER = {
    "validated_leader":  0,
    "validated":         1,
    "robustness":        2,
    "spot_check":        3,
    "cost_sensitivity":  4,
    "rejected":          5,
    "archive_only":      6,
}


def classify_row(label: str, mode: str):
    """Return (exp_name, step, is_new_era)."""
    if not label:
        return None, None, False
    label = str(label)
    is_new_era = label.startswith("exp_") or any(
        label.startswith(p) for p in
        ["cost_sensitivity", "spotcheck", "backtest_8assets",
         "oos_evaluate_8assets", "full_backtest_8assets",
         "5asset_dalio", "5asset_tip", "6asset_", "7asset_", "8asset_"]
    )

    step = "unknown"
    exp_name = label
    for s in ["_s1_backtest", "_s2_optimise", "_s3_walkforward",
              "_s4_oos", "_s5_full", "_spot_full", "_backtest"]:
        if label.endswith(s):
            exp_name = label.replace("exp_", "").replace(s, "")
            step = s.strip("_")
            break
    if label.startswith("exp_"):
        exp_name = exp_name
    elif label.startswith("oos_evaluate_"):
        exp_name = "8asset_manual_v1"
        step = "s4_oos"
    elif label.startswith("full_backtest_"):
        exp_name = "8asset_manual_v1"
        step = "s5_full"
    elif label.startswith("backtest_8assets"):
        exp_name = "8asset_manual_v1"
        step = "s1_backtest"
    else:
        step = "old_era"

    return exp_name, step, is_new_era


def get_oos_calmar(rows_for_exp):
    """Get OOS Calmar from s4_oos rows."""
    for r in rows_for_exp:
        if r["step"] == "s4_oos":
            return r["calmar"] or 0.0
    return None


def assign_group(exp_name, has_oos):
    if exp_name == "6asset_tip_gsg":
        return "validated_leader"
    if "split2018" in exp_name or "split2022" in exp_name:
        return "robustness"
    if "cost_sensitivity" in exp_name:
        return "cost_sensitivity"
    if "spotcheck" in exp_name:
        return "spot_check"
    if exp_name in VALIDATED and has_oos:
        return "validated"
    if has_oos:
        return "validated"
    return "archive_only"


# ── Read source ───────────────────────────────────────────────────────────────

if not os.path.exists(SRC):
    print(f"ERROR: {SRC} not found. Run from project root.")
    exit(1)

src_wb = openpyxl.load_workbook(SRC)
src_ws = src_wb.active

# Read header rows
r1 = [src_ws.cell(1, c).value for c in range(1, src_ws.max_column + 1)]
r2 = [src_ws.cell(2, c).value for c in range(1, src_ws.max_column + 1)]

# Read all data rows
all_rows = []
for row_idx in range(3, src_ws.max_row + 1):
    raw = [src_ws.cell(row_idx, c).value
           for c in range(1, src_ws.max_column + 1)]
    if all(v is None for v in raw):
        continue
    label    = str(raw[1]) if raw[1] else ""
    mode     = str(raw[2]) if raw[2] else ""
    calmar   = raw[19]   # col 20 = AW_R Calmar
    avg_dd   = raw[13]   # col 14 = AvgDD (new metric marker)
    ts       = raw[0]

    exp_name, step, is_new_era = classify_row(label, mode)

    all_rows.append({
        "row_idx":      row_idx,
        "raw":          raw,
        "label":        label,
        "mode":         mode,
        "exp_name":     exp_name,
        "step":         step,
        "is_new_era":   is_new_era,
        "calmar":       calmar,
        "has_full_metrics": avg_dd is not None,
        "ts":           ts,
    })

print(f"Read {len(all_rows)} data rows from {SRC}")

# ── Build curated set ─────────────────────────────────────────────────────────

# Group rows by experiment
by_exp = {}
for r in all_rows:
    e = r["exp_name"]
    if e not in by_exp:
        by_exp[e] = []
    by_exp[e].append(r)

curated_rows = []
archive_only_rows = []

for exp_name, rows in by_exp.items():
    # Old-era: all go to archive only
    if not any(r["is_new_era"] for r in rows):
        archive_only_rows.extend(rows)
        continue

    # Deduplicate: for each step, keep only the last run (highest row_idx)
    by_step = {}
    for r in rows:
        s = r["step"]
        if s not in by_step or r["row_idx"] > by_step[s]["row_idx"]:
            by_step[s] = r

    deduped = list(by_step.values())
    has_oos = "s4_oos" in by_step

    # Decide which steps to include in curated
    for r in deduped:
        step = r["step"]

        # Always archive everything
        # For curated: apply filters
        if step == "old_era" or step == "unknown":
            archive_only_rows.append(r)
            continue

        # s1_backtest: only keep if no s2_optimise exists (IS-only experiment)
        if step == "s1_backtest":
            if "s2_optimise" not in by_step:
                curated_rows.append(r)
            else:
                archive_only_rows.append(r)
            continue

        if step in KEEP_STEPS:
            curated_rows.append(r)
        else:
            archive_only_rows.append(r)

# Sort curated: by group order, then by OOS Calmar desc within group
def sort_key(r):
    exp = r["exp_name"]
    rows_for_exp = by_exp.get(exp, [])
    oos_cal = get_oos_calmar(rows_for_exp) or 0.0
    has_oos = any(x["step"] == "s4_oos" for x in rows_for_exp)
    group = assign_group(exp, has_oos)
    g_order = GROUP_ORDER.get(group, 9)
    # Within experiment, order steps: s2 < s4 < s5 < spot
    step_order = {"s2_optimise": 0, "s4_oos": 1, "s5_full": 2,
                  "spot_full": 3, "s1_backtest": 4}
    s_order = step_order.get(r["step"], 9)
    return (g_order, -oos_cal, exp, s_order)

curated_rows.sort(key=sort_key)

print(f"Curated rows: {len(curated_rows)}")
print(f"Archive-only rows: {len(archive_only_rows)}")


# ── Write output ──────────────────────────────────────────────────────────────

def copy_header_rows(src_ws, tgt_ws, max_col):
    """Copy the two header rows with all styling from source."""
    for r in [1, 2]:
        for c in range(1, max_col + 1):
            src_cell = src_ws.cell(r, c)
            tgt_cell = tgt_ws.cell(r, c)
            tgt_cell.value = src_cell.value
            if src_cell.has_style:
                tgt_cell.font      = src_cell.font.copy()
                tgt_cell.fill      = src_cell.fill.copy()
                tgt_cell.border    = src_cell.border.copy()
                tgt_cell.alignment = src_cell.alignment.copy()

    # Copy merged cells from header
    for merge in src_ws.merged_cells.ranges:
        if merge.min_row <= 2:
            try:
                tgt_ws.merge_cells(str(merge))
            except Exception:
                pass

    # Copy column widths
    for col_letter, col_dim in src_ws.column_dimensions.items():
        tgt_ws.column_dimensions[col_letter].width = col_dim.width


def write_data_rows(rows, tgt_ws, start_row, max_col, src_ws):
    """Write data rows, optionally adding a group separator line."""
    thin   = Side(style="thin", color="444444")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Colour scheme for curated tab
    FILL_LEADER  = PatternFill("solid", fgColor="FFF8E8")  # gold tint
    FILL_VALID   = PatternFill("solid", fgColor="FFFFFF")
    FILL_ROBUST  = PatternFill("solid", fgColor="F0F8FF")  # blue tint
    FILL_SPOT    = PatternFill("solid", fgColor="F8F8F8")
    FILL_COST    = PatternFill("solid", fgColor="F5FFF5")  # green tint
    FILL_ODD     = PatternFill("solid", fgColor="F7F9FB")
    FILL_EVEN    = PatternFill("solid", fgColor="FFFFFF")

    prev_exp = None
    excel_row = start_row

    for i, r in enumerate(rows):
        raw = r["raw"]
        exp = r["exp_name"]

        # Add subtle separator when experiment group changes
        if prev_exp != exp and prev_exp is not None:
            excel_row += 0  # no blank row, just visual border via fill

        # Pick fill
        has_oos = any(x["step"] == "s4_oos" for x in by_exp.get(exp, []))
        group = assign_group(exp, has_oos)

        if group == "validated_leader":
            fill = FILL_LEADER
        elif group == "robustness":
            fill = FILL_ROBUST
        elif group == "cost_sensitivity":
            fill = FILL_COST
        elif group == "spot_check":
            fill = FILL_SPOT
        else:
            fill = FILL_EVEN if i % 2 == 0 else FILL_ODD

        for c in range(1, max_col + 1):
            val = raw[c - 1] if c - 1 < len(raw) else None
            if isinstance(val, float) and val != val:
                val = None
            cell = tgt_ws.cell(excel_row, c, value=val)
            cell.fill   = fill
            cell.border = border
            cell.font   = Font(color="000000", size=9)
            if isinstance(val, (int, float)) and val is not None:
                cell.alignment = Alignment(horizontal="right")
                if isinstance(val, float):
                    cell.number_format = "0.000"
            else:
                cell.alignment = Alignment(horizontal="left")

        prev_exp = exp
        excel_row += 1

    return excel_row


wb_out = Workbook()

# ── TAB 1: Curated ────────────────────────────────────────────────────────────
ws_cur = wb_out.active
ws_cur.title = "Curated"

max_col = src_ws.max_column
copy_header_rows(src_ws, ws_cur, max_col)

# Add a legend row before data
ws_cur.row_dimensions[3].height = 14
legend_texts = {
    1: "★ LEGEND:",
    2: "Gold = validated leader (6asset_tip_gsg) · Blue = robustness tests · "
       "Green = cost sensitivity · White = all other validated experiments",
}
for c, txt in legend_texts.items():
    cell = ws_cur.cell(3, c, value=txt)
    cell.font = Font(italic=True, size=8, color="666666")
    if c == 1:
        cell.font = Font(bold=True, size=8, color="666666")

# Merge legend across all columns
try:
    ws_cur.merge_cells(start_row=3, start_column=2,
                       end_row=3, end_column=max_col)
except Exception:
    pass

write_data_rows(curated_rows, ws_cur, 4, max_col, src_ws)
ws_cur.freeze_panes = "L4"

# ── TAB 2: Archive ────────────────────────────────────────────────────────────
ws_arc = wb_out.create_sheet("Archive (all runs)")
copy_header_rows(src_ws, ws_arc, max_col)

thin   = Side(style="thin", color="444444")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
for i, r in enumerate(all_rows):
    excel_row = i + 3
    raw = r["raw"]
    fill = PatternFill("solid", fgColor="FFFFFF" if i % 2 == 0 else "F7F9FB")
    for c in range(1, max_col + 1):
        val = raw[c - 1] if c - 1 < len(raw) else None
        if isinstance(val, float) and val != val:
            val = None
        cell = ws_arc.cell(excel_row, c, value=val)
        cell.fill   = fill
        cell.border = border
        cell.font   = Font(color="000000", size=9)
        cell.alignment = Alignment(
            horizontal="right" if isinstance(val, (int, float)) and val is not None
            else "left"
        )

ws_arc.freeze_panes = "L3"

# ── Save ──────────────────────────────────────────────────────────────────────
os.makedirs(os.path.dirname(os.path.abspath(OUT)), exist_ok=True)
wb_out.save(OUT)
print(f"\nSaved -> {OUT}")
print(f"  Tab 'Curated':          {len(curated_rows)} rows")
print(f"  Tab 'Archive':          {len(all_rows)} rows")
print(f"\nCurated breakdown:")

seen = set()
for r in curated_rows:
    e = r["exp_name"]
    if e not in seen:
        has_oos = any(x["step"] == "s4_oos" for x in by_exp.get(e, []))
        group = assign_group(e, has_oos)
        oos_cal = get_oos_calmar(by_exp.get(e, [])) or "-"
        print(f"  [{group:<18}] {e:<45} OOS Cal={oos_cal}")
        seen.add(e)