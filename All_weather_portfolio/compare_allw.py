"""
compare_allw.py
===============
Head-to-head comparison of validated All Weather strategies vs the
Bridgewater ALLW ETF (launched March 2025, ticker ALLW on NASDAQ).

All metrics are computed from DAILY prices over the ALLW overlap window,
so MaxDD and Ulcer Index reflect intraday drawdowns, not monthly snapshots.

Fee-adjusted rows apply annualised expense-ratio drag to the gross series:
  ALLW: 0.85% p.a.   DIY strategies: 0.12% p.a. (weighted avg ETF expense)

Outputs
-------
  - Side-by-side performance table (stdout)
  - results/allw_comparison_growth.png
  - results/allw_fee_drag.png              

Usage
-----
  conda run -n allweather python3 compare_allw.py
"""

from __future__ import annotations

import os
import sys
import warnings
from datetime import date, datetime

import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import matplotlib.dates as mdates
import numpy as np
import pandas as pd
import yfinance as yf
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore", category=FutureWarning)
warnings.filterwarnings("ignore", message=".*auto_adjust.*")

os.makedirs("results", exist_ok=True)

# ---------------------------------------------------------------------------
# STRATEGY DEFINITIONS
# ---------------------------------------------------------------------------

# Backtest ETF | Live ETF | Annual fee saving | Status |
#-------------|---------|-------------------|--------|
# SPY → | IVV | 0.03% | ✅ Confirmed identical |
# GLD → | GLDM | 0.30% | ✅ Confirmed identical |
# GSG → | PDBC | 0.16% | ✅ Better contango management |
# QQQ → | QQQM | 0.05% | ✅ Same index |

DATE_START = "2025-03-06"  # ALLW launch date
#DATE_START = "2026-01-01"

#DATE_END   = "2025-06-30"
DATE_END   = date.today().strftime("%Y-%m-%d")

ALLOC_6ASSET_MANUAL_LIVE = {
    "IVV": 0.15, "QQQM": 0.15, "TLT": 0.30,
    "TIP": 0.15, "GLDM": 0.15, "PDBC": 0.10,
}
ALLOC_6ASSET_MANUAL_BACKTEST = {
    "SPY": 0.15, "QQQ": 0.15, "TLT": 0.30,
    "TIP": 0.15, "GLD": 0.15, "GSG": 0.10,
}

ALLOC_6ASSET_RPsplit2020_LIVE = {
    "IVV": 0.13, "QQQM": 0.11, "TLT": 0.19,
    "TIP": 0.33, "GLDM": 0.14, "PDBC": 0.1,
}
ALLOC_6ASSET_RPsplit2020_BACKTEST = {
    "SPY": 0.13, "QQQ": 0.11, "TLT": 0.19,
    "TIP": 0.33, "GLD": 0.14, "GSG": 0.1,
}

ALLOC_6ASSET_RPAVG_LIVE = {
    "IVV": 0.134, "QQQM": 0.103, "TLT": 0.175,
    "TIP": 0.348, "GLDM": 0.142, "PDBC": 0.098,
}
ALLOC_6ASSET_RPAVG_BACKTEST = {
    "SPY": 0.134, "QQQ": 0.103, "TLT": 0.175,
    "TIP": 0.348, "GLD": 0.142, "GSG": 0.098,
}

ALLOC_5ASSET_DALIO_LIVE = {
    "IVV": 0.30, "TLT": 0.40, "IEF": 0.15, "GLDM": 0.075, "PDBC": 0.075,
}
ALLOC_5ASSET_DALIO_BACKTEST = {
    "SPY": 0.30, "TLT": 0.40, "IEF": 0.15, "GLD": 0.075, "GSG": 0.075,
}

ALL_TICKERS   = sorted(set(
    list(ALLOC_6ASSET_MANUAL_LIVE) 
    + list(ALLOC_6ASSET_MANUAL_BACKTEST)
    + list(ALLOC_6ASSET_RPsplit2020_LIVE) 
    + list(ALLOC_6ASSET_RPsplit2020_BACKTEST)
    + list(ALLOC_6ASSET_RPAVG_LIVE) 
    + list(ALLOC_6ASSET_RPAVG_BACKTEST)
    + list(ALLOC_5ASSET_DALIO_LIVE) 
    + list(ALLOC_5ASSET_DALIO_BACKTEST)
))

# SPY and TLT are always needed: SPY as benchmark, TLT for the 60/40 series.
FETCH_TICKERS = sorted(set(ALL_TICKERS + ["ALLW", "SPY", "TLT"]))

ALLW_FEE = 0.0085    # Bridgewater ALLW annual expense ratio
DIY_FEE  = 0.0012    # Weighted avg ETF expense ratio for DIY strategies

IRAN_WAR_DATE = pd.Timestamp("2026-02-28")
WATERMARK     = "github.com/fcastelasimao/quant-learning"

# DARK THEME COLOURS ------------------------------------------------------------

DARK_BG     = "#0d1117"
PANEL_BG    = "#161b22"
GRID_COL    = "#30363d"
TEXT_COL    = "#c9d1d9"
BORDER_COL  = "#30363d"

COLORS = {
    "6asset_manual_live":           "#58a6ff",   
    "6asset_manual_backtest":       "#FF1919",   
    "6asset_rpsplit2020_live":      "#d2a8ff",   
    "6asset_rpsplit2020_backtest":  "#ff3939",   
    "6asset_rpavg_live":            "#d2a8ff",   
    "6asset_rpavg_backtest":        "#58a6ff",   
    "5dalio_live":                  "#3ddc97",   
    "5dalio_backtest":              "#3ddc97",   
    "ALLW":                         "#f0b429",   
    "60/40":                        "#3fb950",   
    "SPY":                          "#f78166",   
    "ALLW_fa":                      "#b08800",   
    "6asset_fa":                    "#1f6feb",   
}

# DATA FETCHING ------------------------------------------------------------------


def _strip_tz(idx: pd.DatetimeIndex) -> pd.DatetimeIndex:
    """Remove timezone info so all indices compare cleanly."""
    if idx.tz is not None:
        return idx.tz_localize(None)
    return idx


def fetch_daily_prices(start: str = DATE_START,
                       end:   str = DATE_END) -> pd.DataFrame:
    """
    Fetch daily adjusted close prices for all comparison tickers.

    Returns a DataFrame indexed by timezone-naive date.
    Raises SystemExit with a clear message if ALLW data is unavailable.
    """
    print(f"Fetching daily prices | {' '.join(FETCH_TICKERS)}")
    print(f"  Period: {start}  →  {end}")

    frames = {}
    failed = []
    for ticker in FETCH_TICKERS:
        try:
            t    = yf.Ticker(ticker)
            hist = t.history(start=start, end=end, auto_adjust=True)
            if hist.empty:
                failed.append(ticker)
                continue
            hist.index = _strip_tz(hist.index)
            frames[ticker] = hist["Close"].rename(ticker)
        except Exception as exc:
            print(f"  ERROR fetching {ticker}: {exc}")
            failed.append(ticker)

    if "ALLW" in failed:
        print(
            "\n" + "=" * 60 + "\n"
            "DATA UNAVAILABLE: ALLW price history could not be retrieved\n"
            "from Yahoo Finance.  ALLW (Bridgewater All Weather ETF) was\n"
            "launched in March 2025.  If yfinance does not yet carry it,\n"
            "try again later or source the data manually from:\n"
            "  https://finance.yahoo.com/quote/ALLW/history\n"
            "=" * 60
        )
        sys.exit(1)

    if failed:
        print(f"  WARNING: could not fetch {failed}; proceeding without them.")

    prices = pd.DataFrame(frames).dropna(how="all")
    prices = prices.ffill()

    print(f"  {len(prices)} trading days loaded"
          f"  ({prices.index[0].date()} → {prices.index[-1].date()})\n")
    return prices


# DAILY PORTFOLIO SERIES (buy-and-hold from day 1) -----------------------------------

def build_daily_series_buy_and_hold(prices: pd.DataFrame,
                       allocation: dict,
                       start_value: float = 100.0) -> pd.Series:
    """
    Compute daily portfolio value using a buy-and-hold strategy
    (initial weights, no rebalancing). Starting value = start_value.

    Returns a Series indexed by date.
    """
    tickers  = [t for t in allocation if t in prices.columns]
    alloc    = {t: allocation[t] for t in tickers}
    # Renormalise in case some tickers are missing
    total_w  = sum(alloc.values())
    alloc    = {t: w / total_w for t, w in alloc.items()}

    first    = prices[tickers].iloc[0]
    shares   = {t: (start_value * w) / float(first[t]) for t, w in alloc.items()}
    daily    = sum(shares[t] * prices[t] for t in tickers)
    return daily.rename("portfolio")

def build_daily_series(prices: pd.DataFrame,
                       allocation: dict,
                       start_value: float = 100.0,
                       rebalance: bool = True) -> pd.Series:
    tickers = [t for t in allocation if t in prices.columns]
    alloc = {t: allocation[t] for t in tickers}
    total_w = sum(alloc.values())
    alloc = {t: w / total_w for t, w in alloc.items()}

    first = prices[tickers].iloc[0]
    shares = {t: (start_value * w) / float(first[t]) for t, w in alloc.items()}

    if not rebalance:
        daily = sum(shares[t] * prices[t] for t in tickers)
        return daily.rename("portfolio")

    # Monthly rebalancing: restore target weights at each month-end
    month_ends = set(prices.resample("ME").last().dropna(how="all").index)
    values = []

    for date, row in prices[tickers].iterrows():
        port_val = sum(shares[t] * float(row[t]) for t in tickers)
        values.append(port_val)

        if date in month_ends:
            for t, w in alloc.items():
                shares[t] = (port_val * w) / float(row[t])

    return pd.Series(values, index=prices[tickers].index, name="portfolio")

def apply_annual_fee(series: pd.Series, annual_fee: float) -> pd.Series:
    """
    Apply a compounding annual expense-ratio drag to a value series.

    Formula: value[n] *= (1 - annual_fee) ^ (n / 252)
    where n = number of trading days since inception.
    """
    n_days   = np.arange(len(series))
    discount = (1.0 - annual_fee) ** (n_days / 252.0)
    return series * discount


# DAILY STATISTICS -------------------------------------------------------------

def _daily_stats(series: pd.Series,
                 label:  str,
                 allocation: dict = None,
                 annualise: int = 252) -> dict:
    """
    Compute all comparison metrics from a daily value series.

    Parameters
    ----------
    series    : daily portfolio value (any starting level)
    label     : display name
    annualise : trading days per year for Sharpe / Sortino (default 252)
    """
    s     = series.dropna()
    if len(s) < 2:
        return {k: float("nan") for k in
                ["label", "total_ret", "cagr", "max_dd", "calmar",
                 "ulcer", "sortino", "vol", "worst_month", "best_month",
                 "n_days"]}

    n_days   = len(s)
    years    = n_days / annualise
    total_r  = (s.iloc[-1] / s.iloc[0] - 1) * 100
    cagr     = ((s.iloc[-1] / s.iloc[0]) ** (1 / years) - 1) * 100

    peak     = s.cummax()
    dd_pct   = ((s - peak) / peak) * 100
    max_dd   = dd_pct.min()

    calmar   = cagr / abs(max_dd) if abs(max_dd) > 1e-6 else 0.0
    ulcer    = float(np.sqrt((dd_pct ** 2).mean()))

    rets     = s.pct_change().dropna()
    down     = rets[rets < 0]
    if len(down) == 0 or down.std() < 1e-10:
        sortino = 0.0
    else:
        sortino = float((rets.mean() / down.std()) * np.sqrt(annualise))

    # Annualised daily volatility (%)
    vol = float(rets.std() * np.sqrt(annualise) * 100)

    # Monthly returns for worst/best month
    monthly_prices = s.resample("ME").last()
    monthly_rets   = monthly_prices.pct_change().dropna() * 100
    worst_month    = float(monthly_rets.min()) if len(monthly_rets) > 0 else float("nan")
    best_month     = float(monthly_rets.max()) if len(monthly_rets) > 0 else float("nan")

    # Allocations
    allocs = " | ".join(f"{t}={w:.1%}" for t, w in allocation.items()) if allocation is not None else ""


    return {
        "label":       label,
        "allocations": allocs,
        "total_ret":   round(total_r,    2),
        "cagr":        round(cagr,       2),
        "max_dd":      round(max_dd,     2),
        "calmar":      round(calmar,     3),
        "ulcer":       round(ulcer,      3),
        "sortino":     round(sortino,    3),
        "vol":         round(vol,        2),
        "worst_month": round(worst_month, 2),
        "best_month":  round(best_month,  2),
        "n_days":      n_days,
    }

# PRINT TABLE --------------------------------------------------------------------------

def print_comparison_table(rows: list[dict],
                            period_label: str) -> None:
    """Print a formatted side-by-side comparison table to stdout."""
    SEP  = "─" * 90
    HDR  = f"{'Strategy':<35} {"Allocations":<75} {'TotRet':>8} {'CAGR':>8} {'MaxDD':>8} "  \
           f"{'Calmar':>8} {'Ulcer':>8} {'Sortino':>9}"

    print()
    print(SEP)
    print(f"  ALLW HEAD-TO-HEAD COMPARISON  ─  {period_label}")
    print(SEP)
    print(HDR)
    print(SEP)

    for r in rows:
        if r is None:
            print()
            continue
        label    = r["label"]
        allocation = r["allocations"]
        tot_r    = f"{r['total_ret']:>+.2f}%"
        cagr     = f"{r['cagr']:>+.2f}%"
        max_dd   = f"{r['max_dd']:.2f}%"
        calmar   = f"{r['calmar']:.3f}"
        ulcer    = f"{r['ulcer']:.3f}"
        sortino  = f"{r['sortino']:+.3f}"
        indent   = "  → " if label.startswith("  ") else ""
        print(f"  {label:<35}  {allocation:<75}  {tot_r:>8} {cagr:>8} {max_dd:>9} " \
              f"{calmar:>8} {ulcer:>8} {sortino:>9}")

    print(SEP)
    print(f"  Metrics computed from daily prices  |  MaxDD = daily peak-to-trough")
    print(f"  Fee-adjusted rows apply annualised expense-ratio drag to gross returns")
    print(f"    ALLW: 0.85% p.a.  |  DIY (6asset / 7asset): 0.12% p.a.")
    print(SEP)
    print()

# DARK THEME HELPER ---------------------------------------------------------------

def _style_ax(ax):
    ax.set_facecolor(PANEL_BG)
    ax.tick_params(colors=TEXT_COL, labelsize=9)
    for sp in ax.spines.values():
        sp.set_color(BORDER_COL)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.yaxis.label.set_color(TEXT_COL)
    ax.xaxis.label.set_color(TEXT_COL)
    ax.title.set_color("white")
    ax.grid(axis="y", color=GRID_COL, alpha=0.45, linewidth=0.7)
    ax.grid(axis="x", color=GRID_COL, alpha=0.25, linewidth=0.4)


def _add_watermark(ax):
    ax.text(
        0.995, 0.012, WATERMARK,
        transform=ax.transAxes,
        ha="right", va="bottom",
        fontsize=7, color="#444c56", alpha=0.8,
        style="italic",
    )

def _save_both(fig, base_name: str):
    """Save chart at Twitter (1200×675) and Reddit (1080×1080) resolutions."""
    #tw_path = os.path.join("results", f"{base_name}_twitter.png")
    #rd_path = os.path.join("results", f"{base_name}_reddit.png")
    #just save one
    path = os.path.join("results", f"{base_name}.png")

    """
    # Twitter: 16:9 -- save at current fig dimensions then resize via dpi
    fig.set_size_inches(12.0, 6.75)
    plt.savefig(tw_path, dpi=100, bbox_inches="tight",
                facecolor=fig.get_facecolor())
    print(f"  Saved → {tw_path}  (1200×675)")

    # Reddit square: 1:1 -- pad with equal margins
    fig.set_size_inches(10.8, 10.8)
    plt.savefig(rd_path, dpi=100, bbox_inches="tight",
                facecolor=fig.get_facecolor())
    print(f"  Saved → {rd_path}  (1080×1080)")
    """
    plt.savefig(path, dpi=100, bbox_inches="tight",
                facecolor=fig.get_facecolor())
    print(f"  Saved → {path}")


# CHART 1: GROWTH CHART ----------------------------------------------------------------

def plot_growth_chart(daily_series: dict[str, pd.Series]) -> None:
    """
    Cumulative growth chart: $10,000 invested on ALLW launch date.

    Plots all series rebased to $10,000 on the ALLW start date.
    Annotates terminal value for each line and the Iran war start.
    Dark theme. Watermark bottom-right.
    Saved as allw_comparison_growth_twitter.png / _reddit.png.
    """
    print("Building growth chart ...")

    allw_start = daily_series["ALLW"].first_valid_index()

    fig, ax = plt.subplots(figsize=(12, 6.75))
    fig.patch.set_facecolor(DARK_BG)
    _style_ax(ax)

    plot_order = [
        #("6asset_manual_live",     COLORS["6asset_manual_live"],  2.2, "--",  "6asset_manual_live "),
        ("6asset_manual_backtest",     COLORS["6asset_manual_backtest"],  1.5, "--",  "6asset_manual_backtest "),
        #("6asset_rpsplit2020_live", COLORS["6asset_rpsplit2020_live"],  2.2, "--", "6asset_rpsplit2020_live "),
        #("6asset_rpsplit2020_backtest", COLORS["6asset_rpsplit2020_backtest"],  1.5, "--", "6asset_rpsplit2020_backtest "),
        #("6asset_ravg_live", COLORS["6asset_ravg_live"],  2.2, "-", "6asset_ravg_live "),
        ("6asset_rpavg_backtest", COLORS["6asset_rpavg_backtest"], 2.2, "-", "6asset_rpavg_backtest"),
        #("5asset_dalio_live",       COLORS["5dalio_live"],  1.5, "-.", "5asset_dalio (Dalio classic)"),
        #("5asset_dalio_backtest",       COLORS["5dalio_backtest"],  1.5, "-.", "5asset_dalio (Dalio classic)"),
        ("ALLW",               COLORS["ALLW"],    2.2, "-",  "ALLW (Bridgewater, 0.85% fee)"),
        ("SPY",                COLORS["SPY"],     1.5, ":",  "SPY (S&P 500)"),
        #("60/40",              COLORS["60/40"],   1.5, "-.", "60/40 (SPY/TLT)"),
    ]

    for key, color, lw, ls, label in plot_order:
        if key not in daily_series:
            continue
        raw = daily_series[key].dropna()
        s   = raw.loc[allw_start:]
        s   = s / s.iloc[0] * 10_000

        ax.plot(s.index, s.values, color=color, lw=lw, linestyle=ls,
                label=label, alpha=0.92)

        # Terminal value annotation at right edge
        final_val  = s.iloc[-1]
        final_date = s.index[-1]
        ax.annotate(
            f"${final_val:,.0f}",
            xy=(final_date, final_val),
            xytext=(8, 0),
            textcoords="offset points",
            color=color,
            fontsize=9,
            fontweight="bold",
            va="center",
        )

    # --- Iran war annotation (range-check, not exact index membership) ---
    ref = daily_series.get("SPY", daily_series.get("ALLW"))
    if ref is not None:
        s_ref = ref.loc[allw_start:]
        if s_ref.index[0] <= IRAN_WAR_DATE <= s_ref.index[-1]:
            ax.axvline(IRAN_WAR_DATE, color="#ff7b72", lw=1.2,
                       linestyle="--", alpha=0.75)
            ymin, ymax = ax.get_ylim()
            ax.text(
                IRAN_WAR_DATE, ymin + (ymax - ymin) * 0.04,
                " Iran war\n starts",
                color="#ff7b72", fontsize=8.5, va="bottom",
                style="italic", alpha=0.85,
            )

    ax.set_title(
        "$10,000 invested in March 2025 — where are you today?",
        fontsize=12, pad=12, color="white",
    )
    ax.set_ylabel("Portfolio Value ($)", fontsize=10)
    ax.yaxis.set_major_formatter(
        mticker.FuncFormatter(lambda x, _: f"${x:,.0f}"))
    ax.xaxis.set_major_formatter(mdates.DateFormatter("%b %Y"))
    ax.xaxis.set_major_locator(mdates.MonthLocator(interval=2))
    plt.setp(ax.xaxis.get_majorticklabels(), rotation=25, ha="right")

    ax.legend(
        fontsize=9, facecolor="#21262d", edgecolor=BORDER_COL,
        labelcolor=TEXT_COL, loc="upper left", framealpha=0.9,
    )

    _add_watermark(ax)
    plt.tight_layout(pad=1.5)
    _save_both(fig, f"{date.today().strftime('%Y-%m-%d')}_allw_comparison_growth_from{DATE_START}_to{DATE_END}")
    plt.close(fig)

# CHART 2: CUMULATIVE FEE DRAG PROJECTION -----------------------------------------

def plot_fee_drag_chart() -> None:
    """
    Projected cost of fees on a $100k portfolio over 30 years.

    Assumes 7% gross annual return.
    Two lines: ALLW (0.85% p.a.) vs DIY (0.12% p.a.)
    Annotated dollar amounts at 10 / 20 / 30 year marks.
    Dark theme. Watermark bottom-right.
    """
    print("Building fee drag chart ...")

    STARTING_VALUE = 100_000
    GROSS_RETURN   = 0.07
    YEARS          = np.linspace(0, 30, 1000)
    MARKS          = [10, 20, 30]

    gross     = STARTING_VALUE * (1 + GROSS_RETURN) ** YEARS
    allw_net  = STARTING_VALUE * (1 + GROSS_RETURN - ALLW_FEE) ** YEARS
    diy_net   = STARTING_VALUE * (1 + GROSS_RETURN - DIY_FEE)  ** YEARS

    fig, ax = plt.subplots(figsize=(12, 6.75))
    fig.patch.set_facecolor(DARK_BG)
    _style_ax(ax)

    ax.plot(YEARS, gross,    color="#8b949e", lw=1.5, linestyle=":",
            label="Gross (no fees) — 7.00% p.a.", alpha=0.7)
    ax.plot(YEARS, diy_net,  color=COLORS["6asset_manual_live"], lw=2.5,
            label=f"DIY all-weather — 0.12% p.a. fee (net {GROSS_RETURN - DIY_FEE:.2%})")
    ax.plot(YEARS, allw_net, color=COLORS["ALLW"], lw=2.5,
            label=f"ALLW ETF        — 0.85% p.a. fee (net {GROSS_RETURN - ALLW_FEE:.2%})")

    # Fill the fee drag gap
    ax.fill_between(YEARS, diy_net, allw_net,
                    color="#f0b429", alpha=0.08, label="Fee drag gap")

    # --- Annotate 10 / 20 / 30 year marks ---
    for yr in MARKS:
        idx    = np.searchsorted(YEARS, yr)
        diy_v  = diy_net[idx]
        allw_v = allw_net[idx]
        gap    = diy_v - allw_v

        # Dots
        ax.plot(yr, diy_v,  "o", color=COLORS["6asset_manual_live"],  ms=7, zorder=5)
        ax.plot(yr, allw_v, "o", color=COLORS["ALLW"],    ms=7, zorder=5)

        # Dollar labels — DIY above, ALLW below
        ax.annotate(
            f"${diy_v/1e3:.0f}k",
            xy=(yr, diy_v), xytext=(yr + 0.4, diy_v * 1.035),
            color=COLORS["6asset_manual_live"], fontsize=9, fontweight="bold",
            arrowprops=dict(arrowstyle="-", color=COLORS["6asset_manual_live"], lw=0.7),
        )
        ax.annotate(
            f"${allw_v/1e3:.0f}k",
            xy=(yr, allw_v), xytext=(yr + 0.4, allw_v * 0.93),
            color=COLORS["ALLW"], fontsize=9, fontweight="bold",
            arrowprops=dict(arrowstyle="-", color=COLORS["ALLW"], lw=0.7),
        )
        # Gap label centred between the two
        ax.text(
            yr - 0.6, (diy_v + allw_v) / 2,
            f"−${gap/1e3:.0f}k\ngap",
            color="#8b949e", fontsize=7.5, ha="right", va="center",
            style="italic",
        )

    ax.set_title(
        "Cumulative Fee Drag — ALLW (0.85%) vs DIY All Weather (0.12%)\n"
        f"Starting value: ${STARTING_VALUE:,.0f}   Assumed gross return: {GROSS_RETURN:.0%} p.a.",
        fontsize=11, pad=12, color="white",
    )
    ax.set_xlabel("Years", fontsize=10)
    ax.set_ylabel("Portfolio Value ($)", fontsize=10)
    ax.set_xlim(0, 31)
    ax.yaxis.set_major_formatter(
        mticker.FuncFormatter(lambda x, _: f"${x/1e3:,.0f}k"))

    legend = ax.legend(
        fontsize=9.5, facecolor="#21262d", edgecolor=BORDER_COL,
        labelcolor=TEXT_COL, loc="upper left", framealpha=0.92,
    )

    _add_watermark(ax)
    plt.tight_layout(pad=1.5)
    _save_both(fig, f"{date.today().strftime('%Y-%m-%d')}_allw_fee_drag_from{DATE_START}_to{DATE_END}")
    plt.close(fig)


# EXCEL EXPORT ---------------------------------------------------------------------

# Column definitions: (header, dict_key, col_width, number_format | None)
_EXCEL_COLS = [
    ("Strategy",      "label",       25,  None),
    ("Allocations",   "allocations", 70,  None),
    ("Total Ret (%)", "total_ret",   13,  "0.0"),
    ("CAGR (%)",      "cagr",        11,  "0.0"),
    ("Max DD (%)",    "max_dd",      11,  "0.0"),
    ("Calmar",        "calmar",      10,  "0.000"),
    ("Ulcer",         "ulcer",       10,  "0.000"),
    ("Sortino",       "sortino",     10,  "0.000"),
    ("Vol (%)",       "vol",         10,  "0.0"),
    ("Worst Mo (%)",  "worst_month", 13,  "0.0"),
    ("Best Mo (%)",   "best_month",  13,  "0.0"),
]

# Row background colours per strategy group (openpyxl expects no leading #)
_GROUP_COLOURS = {
    "6asset_manual_live": "0D2137",
    "6asset_manual_backtest": "1A0D37",
    "6asset_rpsplit2020_live": "0D2A1A",
    "6asset_rpsplit2020_backtest": "2A1F00",
    "6asset_ravg_live": "0D2A1A",
    "6asset_ravg_backtest": "2A1F00",
    "5asset_dalio_live": "2A0D0D",
    "5asset_dalio_backtest": "0D2A0D",
}
_DEFAULT_ROW_COLOUR = "161B22"
_SPACER_COLOUR      = "21262D"


def _xfill(hex_colour: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_colour.lstrip("#").upper())


def _xfont(bold: bool = False, italic: bool = False,
           colour: str = "C9D1D9", size: int = 10) -> Font:
    return Font(bold=bold, italic=italic,
                color=colour.lstrip("#").upper(), size=size)


def _xborder(bottom_colour: str = "444444") -> Border:
    thin   = Side(style="thin", color="444444")
    bottom = Side(style="thin", color=bottom_colour.lstrip("#").upper())
    return Border(left=thin, right=thin, top=thin, bottom=bottom)


def _row_colour(label: str, current_group: str) -> str:
    """Return the background hex for a data row (no leading #)."""
    for prefix, colour in _GROUP_COLOURS.items():
        if label.startswith(prefix):
            return colour
    if label.startswith("  "):   # fee-adjusted row: inherit parent group
        return current_group
    return _DEFAULT_ROW_COLOUR


def save_comparison_excel(rows: list, period_label: str) -> None:
    """
    Write the comparison table to results/allw_comparison_<YYYYMMDD>.xlsx.

    Completely independent of master_log.xlsx — never reads from or writes
    to that file.
    """
    today    = datetime.now()
    filename = f"{date.today().strftime('%Y-%m-%d')}_allw_comparison_from{DATE_START}_to{DATE_END}.xlsx"
    out_path = os.path.join("results", filename)

    n_cols       = len(_EXCEL_COLS)
    center_align = Alignment(horizontal="center", vertical="center",
                             wrap_text=True)
    right_align  = Alignment(horizontal="right",  vertical="center")
    left_align   = Alignment(horizontal="left",   vertical="center")

    wb = Workbook()
    ws = wb.active
    ws.title = "ALLW Comparison"

    # ── Row 1: period label merged across all columns ────────────────────────
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1,   end_column=n_cols)
    cell           = ws.cell(row=1, column=1, value=period_label)
    cell.fill      = _xfill("0D1117")
    cell.font      = _xfont(bold=True, colour="FFFFFF", size=11)
    cell.alignment = center_align

    # ── Row 2: column headers ────────────────────────────────────────────────
    for col_i, (header, _, width, _) in enumerate(_EXCEL_COLS, start=1):
        cell           = ws.cell(row=2, column=col_i, value=header)
        cell.fill      = _xfill("161B22")
        cell.font      = _xfont(bold=True, colour="FFFFFF", size=10)
        cell.alignment = center_align
        cell.border    = _xborder()
        ws.column_dimensions[get_column_letter(col_i)].width = width

    # ── Data rows ────────────────────────────────────────────────────────────
    current_group = _DEFAULT_ROW_COLOUR
    excel_row     = 3

    for i, row in enumerate(rows):
        if row is None:
            # Spacer: fill every cell, no border
            for col_i in range(1, n_cols + 1):
                ws.cell(row=excel_row, column=col_i).fill = _xfill(_SPACER_COLOUR)
            excel_row += 1
            continue

        label       = row.get("label", "")
        row_colour  = _row_colour(label, current_group)
        is_fee_row  = label.startswith("  ")
        if not is_fee_row:
            current_group = row_colour

        # Thick bottom border when the next item is None or end-of-list
        next_item      = rows[i + 1] if i + 1 < len(rows) else None
        is_group_end   = (next_item is None)
        bottom_col     = "888888" if is_group_end else "444444"

        for col_i, (_, key, _, num_fmt) in enumerate(_EXCEL_COLS, start=1):
            value = row.get(key, float("nan"))
            # Replace nan with empty string
            if isinstance(value, float) and pd.isna(value):
                value = ""

            cell        = ws.cell(row=excel_row, column=col_i, value=value)
            cell.fill   = _xfill(row_colour)
            cell.border = _xborder(bottom_colour=bottom_col)

            if col_i == 1:   # Strategy column
                cell.font      = _xfont(bold=not is_fee_row,
                                        italic=is_fee_row,
                                        colour="C9D1D9")
                cell.alignment = left_align
            else:
                cell.font      = _xfont(colour="C9D1D9")
                cell.alignment = right_align
                if num_fmt and isinstance(value, (int, float)):
                    cell.number_format = num_fmt

        excel_row += 1

    # ── Footer (blank gap, then three lines) ─────────────────────────────────
    excel_row += 1   # blank gap
    footer_lines = [
        f"Generated: {today.strftime('%Y-%m-%d %H:%M')}",
        f"Period: {period_label}",
        ("Note: All metrics computed from daily prices. ALLW data from "
         "Yahoo Finance. DIY strategies use buy-and-hold from ALLW launch "
         "date. Fee-adjusted rows apply annual expense ratio drag compounded "
         "daily. Rf = 0 assumed for Sortino."),
    ]
    small_grey = _xfont(colour="8B949E", size=9)
    for line in footer_lines:
        ws.merge_cells(start_row=excel_row, start_column=1,
                       end_row=excel_row,   end_column=n_cols)
        cell      = ws.cell(row=excel_row, column=1, value=line)
        cell.font = small_grey
        excel_row += 1

    # ── Freeze at B3 ─────────────────────────────────────────────────────────
    ws.freeze_panes = "B3"

    wb.save(out_path)
    print(f"  Excel saved → {out_path}")


# MAIN ----------------------------------------------------------------------

def main():
    # ── 1. Fetch data ────────────────────────────────────────────────────────
    prices = fetch_daily_prices(start=DATE_START, end=DATE_END)

    # Align all series to the ALLW window
    allw_start = prices["ALLW"].first_valid_index()
    prices     = prices.loc[allw_start:]
    period_label = (f"{prices.index[0].date()} → {prices.index[-1].date()}  "
                    f"({len(prices)} trading days,  "
                    f"{len(prices)/252:.2f} yrs)  ")

    # ── 2. Build daily portfolio value series ────────────────────────────────

    STRATEGIES = [
        ("6asset_manual_live", ALLOC_6ASSET_MANUAL_LIVE, DIY_FEE, "6asset_manual_live (gross)"),
        ("6asset_manual_backtest", ALLOC_6ASSET_MANUAL_BACKTEST, DIY_FEE, "6asset_manual_backtest (gross)"),
        ("6asset_rpsplit2020_live", ALLOC_6ASSET_RPsplit2020_LIVE, DIY_FEE, "6asset_rpsplit2020_live (gross)"),
        ("6asset_rpsplit2020_backtest", ALLOC_6ASSET_RPsplit2020_BACKTEST, DIY_FEE, "6asset_rpsplit2020_backtest (gross)"),
        ("6asset_rpavg_live", ALLOC_6ASSET_RPAVG_LIVE, DIY_FEE, "6asset_rpavg_live (gross)"),
        ("6asset_rpavg_backtest", ALLOC_6ASSET_RPAVG_BACKTEST, DIY_FEE, "6asset_rpavg_backtest (gross)"),
        ("5asset_dalio_live", ALLOC_5ASSET_DALIO_LIVE, DIY_FEE, "5asset_dalio_live (gross)"),
        ("5asset_dalio_backtest", ALLOC_5ASSET_DALIO_BACKTEST, DIY_FEE, "5asset_dalio_backtest (gross)"),
    ]

    series = {}
    rows = []
    for key, alloc, fee, label in STRATEGIES:
        s = build_daily_series(prices, alloc, rebalance=True)
        s_fa = apply_annual_fee(s, fee)
        series[key] = s
        rows.append(_daily_stats(s, label, allocation=alloc))
        rows.append(_daily_stats(s_fa, f"  → fee-adj {fee:.2%} p.a."))
        rows.append(None)  # spacer

    s_allw   = prices["ALLW"] / prices["ALLW"].iloc[0] * 100.0
    s_spy    = prices["SPY"]  / prices["SPY"].iloc[0]  * 100.0
    # 60/40: 60% SPY, 40% TLT, buy-and-hold
    if "TLT" in prices.columns:
        spy_sh   = 0.60 / float(prices["SPY"].iloc[0])
        tlt_sh   = 0.40 / float(prices["TLT"].iloc[0])
        s_6040   = (spy_sh * prices["SPY"] + tlt_sh * prices["TLT"]) * 100.0 / 1.0
        # Re-base to 100
        s_6040   = s_6040 / s_6040.iloc[0] * 100.0
    else:
        s_6040 = None

    # Fee-adjusted versions of ALLW
    s_allw_fa = apply_annual_fee(s_allw,    ALLW_FEE)

    # Add ALLW and benchmark rows at the end for better visibility
    rows.append(_daily_stats(s_allw,      "ALLW  (Bridgewater, gross)"))
    rows.append(_daily_stats(s_allw_fa,   "  → fee-adj 0.85% p.a.", allocation = None)) 
    rows.append(None)  # spacer
    rows.append(_daily_stats(s_spy,       "SPY  (benchmark)", allocation = None))
    if s_6040 is not None:
        rows.append(_daily_stats(s_6040, "60/40  (SPY 60% / TLT 40%)"))

    # ── 3. Print table ───────────────────────────────────────────────────────
    print_comparison_table(rows, period_label)

    # ── 3b. Excel export ─────────────────────────────────────────────────────
    save_comparison_excel(rows, period_label)

    # ── 4. Chart 1: Growth chart ──────────────────────────────────────────────
    daily_series = {asset: s for asset, s in series.items()}
    daily_series["ALLW"] = s_allw
    daily_series["SPY"]  = s_spy
    
    if s_6040 is not None:
        daily_series["60/40"] = s_6040

    plot_growth_chart(daily_series)

    # ── 6. Chart 2: Fee drag projection ──────────────────────────────────────
    #plot_fee_drag_chart()

    print("\nDone.")


if __name__ == "__main__":
    main()
