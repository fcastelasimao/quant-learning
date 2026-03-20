"""
export.py
=========
Everything that writes files to disk.

Contains:
  - make_results_dir      create timestamped folder for this run
  - save_run_config       save all parameters as JSON for reproducibility
  - export_results        save backtest CSV, stats CSV, allocation CSV
  - build_log_row         build one summary row from a list of StrategyStats
  - append_to_master_log  append that row to results/master_log.xlsx
  - print_header          print a section divider to terminal
  - print_rebalancing     print formatted rebalancing instructions
  - print_stats           print formatted performance statistics

Master log is now an Excel file (master_log.xlsx) with grouped/merged
headers so strategy names appear once above their metrics columns.
Requires openpyxl >= 3.1.0.
"""

from __future__ import annotations

import json
import os
from datetime import datetime

import pandas as pd
from openpyxl.styles import (Alignment, Font, PatternFill, Border, Side)
from openpyxl.utils import get_column_letter

from backtest import StrategyStats
import config

import sys

# ===========================================================================
# RESULTS DIRECTORY
# ===========================================================================

class _Tee:
    """Write to both stdout and a log file simultaneously."""
    def __init__(self, filepath):
        self._file = open(filepath, "w", buffering=1)
        self._stdout = sys.stdout

    def write(self, data):
        self._stdout.write(data)
        self._file.write(data)

    def flush(self):
        self._stdout.flush()
        self._file.flush()

    def close(self):
        sys.stdout = self._stdout
        self._file.close()

def start_run_log(results_dir: str):
    """Redirect stdout to both terminal and run_log.txt."""
    path = os.path.join(results_dir, "run_log.txt")
    tee  = _Tee(path)
    sys.stdout = tee
    return tee

def stop_run_log(tee):
    """Restore stdout and close the log file."""
    tee.close()

def make_results_dir(label: str) -> str:
    """
    Create a timestamped folder inside 'results/' for this run.
    Format: results/YYYY-MM-DD_HH-MM-SS_<label>/

    Every run gets its own folder so results are never overwritten.
    Returns the path to the created folder.
    """
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    folder    = os.path.join("results", f"{timestamp}_{label}")
    os.makedirs(folder, exist_ok=True)
    return folder


# ===========================================================================
# RUN CONFIG
# ===========================================================================

def save_run_config(allocation: dict, results_dir: str, label: str):
    """
    Save all user parameters and the final allocation to run_config.json.

    This file contains everything needed to reproduce the run exactly:
    copy these values back into config.py and re-run main.py.
    """
    config_data = {
        "run_label":               label,
        "timestamp":               datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "backtest_start":          config.BACKTEST_START,
        "backtest_end":            config.BACKTEST_END,
        "oos_start":               config.OOS_START,
        "pricing_model":           config.PRICING_MODEL,
        "transaction_cost_pct":    config.TRANSACTION_COST_PCT,
        "tax_drag_pct":            config.TAX_DRAG_PCT,
        "initial_portfolio_value": config.INITIAL_PORTFOLIO_VALUE,
        "rebalance_threshold":     config.REBALANCE_THRESHOLD,
        "data_frequency":          config.DATA_FREQUENCY,
        "sharpe_annualisation":    config.SHARPE_ANNUALISATION,
        "benchmark_ticker":        config.BENCHMARK_TICKER,
        "optimiser": {
            "method":       config.OPT_METHOD if config.RUN_MODE == "optimise" else "not run",
            "min_weight":    config.OPT_MIN_WEIGHT,
            "max_weight":    config.OPT_MAX_WEIGHT,
            "min_cagr":      config.OPT_MIN_CAGR,
            "n_trials":      config.OPT_N_TRIALS,
            "random_seed":   config.OPT_RANDOM_SEED,
        },
        "pareto": {
            "cagr_range": list(config.PARETO_CAGR_RANGE) if config.RUN_MODE == "pareto" else [],
        },
        "walk_forward": {
            "train_years": config.WF_TRAIN_YEARS,
            "test_years":  config.WF_TEST_YEARS,
            "step_years":  config.WF_STEP_YEARS,
            "opt_method":  config.WF_OPT_METHOD,
        },
        "run_mode": config.RUN_MODE,
        "target_allocation": allocation,
    }

    path = os.path.join(results_dir, "run_config.json")
    with open(path, "w") as f:
        json.dump(config_data, f, indent=2)


# ===========================================================================
# RESULTS EXPORT
# ===========================================================================

def export_results(backtest: pd.DataFrame,
                   instructions: pd.DataFrame,
                   stats_list: list[StrategyStats],
                   allocation: dict,
                   results_dir: str,
                   label: str):
    """
    Export all results to results_dir:
      backtest_history.csv         -- monthly portfolio values
      rebalancing_instructions.csv
      stats.csv                    -- CAGR, drawdown, Sharpe, Calmar per strategy
      allocation.csv               -- weights used in this run
      run_config.json              -- all parameters for reproduction
    """
    stats_rows = []
    for s in stats_list:
        stats_rows.extend([
            {"Strategy": s.name, "Metric": "Period (years)",            "Value": s.period_years},
            {"Strategy": s.name, "Metric": "CAGR (%)",                  "Value": s.cagr},
            {"Strategy": s.name, "Metric": "Max Drawdown (%)",          "Value": s.max_drawdown},
            {"Strategy": s.name, "Metric": "Avg Drawdown (%)",          "Value": s.avg_drawdown},
            {"Strategy": s.name, "Metric": "Max DD Duration (months)",  "Value": s.max_dd_duration},
            {"Strategy": s.name, "Metric": "Avg Recovery (months)",     "Value": s.avg_recovery_time},
            {"Strategy": s.name, "Metric": "Ulcer Index",               "Value": s.ulcer_index},
            {"Strategy": s.name, "Metric": "Sharpe Ratio",              "Value": s.sharpe},
            {"Strategy": s.name, "Metric": "Sortino Ratio",             "Value": s.sortino},
            {"Strategy": s.name, "Metric": "Calmar Ratio",              "Value": s.calmar},
            {"Strategy": s.name, "Metric": "Final Value ($)",           "Value": s.final_value},
        ])

    pd.DataFrame(stats_rows).to_csv(
        os.path.join(results_dir, "stats.csv"), index=False)

    backtest.to_csv(
        os.path.join(results_dir, "backtest_history.csv"))

    instructions.to_csv(
        os.path.join(results_dir, "rebalancing_instructions.csv"), index=False)

    pd.DataFrame([
        {"Ticker": t, "Weight": w, "Weight (%)": f"{w:.1%}"}
        for t, w in allocation.items()
    ]).to_csv(os.path.join(results_dir, "allocation.csv"), index=False)

    save_run_config(allocation, results_dir, label)

    print(f"  backtest_history.csv")
    print(f"  rebalancing_instructions.csv")
    print(f"  stats.csv")
    print(f"  allocation.csv")
    print(f"  run_config.json")


# ===========================================================================
# MASTER LOG -- EXCEL
# ===========================================================================

# Column definitions for the master log.
# META_COLS   : columns that appear once per run, before the strategy groups
# METRIC_COLS : per-strategy metric columns that are grouped under a header
# STRATEGY_NAMES : the three strategy labels (must match StrategyStats.name)

META_COLS = [
    "Timestamp",
    "Label",
    "Run Mode",
    "Backtest Start",
    "Backtest End",
    "OOS Start",
    "Pricing Model",
    "Tx Cost %",
    "Tax Drag %",
    "Data Freq",
    "Tickers",
]

METRIC_COLS = [
    "CAGR (%)",
    "Max_DD (%)",
    "Avg_DD (%)",
    "Max_DD_Dur",
    "Avg_Rec",
    "Ulcer",
    "Sharpe",
    "Sortino",
    "Calmar",
    "Fin_Val ($)",
]

STRATEGY_NAMES = ["AW_R", "B&H_AW", "SPY", "60/40"]

# Colours for the strategy header groups (subtle dark-friendly palette)
STRATEGY_COLOURS = {
    "AW_R":   "1F4E79",   # dark blue
    "B&H_AW": "7B3F00",   # dark amber/brown
    "SPY":    "3D1A1A",   # dark red
    "60/40":  "1A4731",   # dark green
}

HEADER_FONT_COLOUR = "FFFFFF"   # white text on coloured headers


def _all_flat_columns() -> list[str]:
    """Return the full ordered list of flat column names used in the data rows."""
    cols = list(META_COLS)
    for strategy in STRATEGY_NAMES:
        for metric in METRIC_COLS:
            cols.append(f"{strategy}_{metric}")
    cols.append("Results Folder")
    return cols


def build_log_row(results_dir: str,
                  stats_list: list[StrategyStats],
                  weights: dict,
                  label: str) -> dict:
    """Build one flat dict representing a single run for the master log."""
    row = {
        "Timestamp":      datetime.now().strftime("%Y-%m-%d %H:%M"),
        "Label":          label,
        "Run Mode":       config.RUN_MODE,
        "Backtest Start": config.BACKTEST_START,
        "Backtest End":   config.BACKTEST_END,
        "OOS Start":      config.OOS_START,
        "Pricing Model":  config.PRICING_MODEL,
        "Tx Cost %":      config.TRANSACTION_COST_PCT,
        "Tax Drag %":     config.TAX_DRAG_PCT,
        "Data Freq":      config.DATA_FREQUENCY,
        "Tickers":        " | ".join(f"{t}={w:.1%}" for t, w in weights.items()),
    }
    for s in stats_list:
        row[f"{s.name}_CAGR (%)"]    = s.cagr
        row[f"{s.name}_Max_DD (%)"]  = s.max_drawdown
        row[f"{s.name}_Avg_DD (%)"]  = s.avg_drawdown
        row[f"{s.name}_Max_DD_Dur"]  = s.max_dd_duration
        row[f"{s.name}_Avg_Rec"]     = s.avg_recovery_time
        row[f"{s.name}_Ulcer"]       = s.ulcer_index
        row[f"{s.name}_Sharpe"]      = s.sharpe
        row[f"{s.name}_Sortino"]     = s.sortino
        row[f"{s.name}_Calmar"]      = s.calmar
        row[f"{s.name}_Fin_Val ($)"] = s.final_value

    row["Results Folder"] = results_dir
    return row


def _write_excel_log(log_path: str, rows: list[dict]):
    """
    Write all rows to master_log.xlsx with two header rows:
      Row 1: merged strategy group headers  (AW_R | B&H_AW | SPY)
      Row 2: individual metric headers      (CAGR, Max_DD, Sharpe, ...)
      Row 3+: data

    Meta columns (Timestamp, Label, etc.) span both header rows via merge.
    """
    flat_cols = _all_flat_columns()

    # Build a plain DataFrame with flat column names for the data rows
    df = pd.DataFrame(rows, columns=flat_cols)

    # Write using openpyxl directly so we can add the two-row header
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Master Log"

    # --- Build header row 1 (group labels) and row 2 (metric labels) ---
    # We'll track which Excel column index each flat column sits in
    col_index = {}   # flat_col_name -> excel column number (1-based)

    excel_col = 1

    # Meta columns: merge row 1 and row 2 vertically
    thin = Side(style="thin", color="444444")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    meta_style = PatternFill("solid", fgColor="1A1A2E")
    meta_font  = Font(bold=True, color="C9D1D9")
    center     = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_name in META_COLS:
        cell1 = ws.cell(row=1, column=excel_col, value=col_name)
        cell2 = ws.cell(row=2, column=excel_col, value="")
        ws.merge_cells(start_row=1, start_column=excel_col,
                       end_row=2,   end_column=excel_col)
        cell1.fill      = meta_style
        cell1.font      = meta_font
        cell1.alignment = center
        cell1.border    = border
        col_index[col_name] = excel_col
        excel_col += 1

    # Strategy groups: row 1 = merged group label, row 2 = individual metrics
    for strategy in STRATEGY_NAMES:
        group_start = excel_col
        colour      = STRATEGY_COLOURS.get(strategy, "333333")
        group_fill  = PatternFill("solid", fgColor=colour)
        group_font  = Font(bold=True, color=HEADER_FONT_COLOUR)
        metric_fill = PatternFill("solid", fgColor="161B22")
        metric_font = Font(bold=False, color="C9D1D9")

        for metric in METRIC_COLS:
            flat_name = f"{strategy}_{metric}"
            col_index[flat_name] = excel_col

            # Row 2: metric name
            cell2 = ws.cell(row=2, column=excel_col, value=metric)
            cell2.fill      = metric_fill
            cell2.font      = metric_font
            cell2.alignment = center
            cell2.border    = border

            excel_col += 1

        # Row 1: merged group label spanning all metrics for this strategy
        group_end = excel_col - 1
        cell1 = ws.cell(row=1, column=group_start, value=strategy)
        ws.merge_cells(start_row=1, start_column=group_start,
                       end_row=1,   end_column=group_end)
        cell1.fill      = group_fill
        cell1.font      = group_font
        cell1.alignment = center
        cell1.border    = border

    # Results Folder: merge rows 1+2 like meta cols
    rf_col = excel_col
    cell1  = ws.cell(row=1, column=rf_col, value="Results Folder")
    ws.merge_cells(start_row=1, start_column=rf_col,
                   end_row=2,   end_column=rf_col)
    cell1.fill      = meta_style
    cell1.font      = meta_font
    cell1.alignment = center
    cell1.border    = border
    col_index["Results Folder"] = rf_col

    # --- Write data rows ---
    data_fill_even = PatternFill("solid", fgColor="FFFFFF")
    data_fill_odd  = PatternFill("solid", fgColor="F2F2F2")
    data_font      = Font(color="000000")
    num_align      = Alignment(horizontal="right")
    left_align     = Alignment(horizontal="left")

    for row_idx, row_data in enumerate(rows):
        excel_row = row_idx + 3   # rows 1 and 2 are headers
        fill      = data_fill_even if row_idx % 2 == 0 else data_fill_odd

        for flat_col, ec in col_index.items():
            value = row_data.get(flat_col, "")
            cell  = ws.cell(row=excel_row, column=ec, value=value)
            cell.fill   = fill
            cell.font   = data_font
            cell.border = border
            # Right-align numbers, left-align text
            if isinstance(value, (int, float)):
                cell.alignment = num_align
                # Format floats to 3 decimal places
                if isinstance(value, float):
                    cell.number_format = "0.000"
            else:
                cell.alignment = left_align

    # --- Column widths ---
    ws.column_dimensions[get_column_letter(col_index["Timestamp"])].width      = 16
    ws.column_dimensions[get_column_letter(col_index["Label"])].width          = 30
    ws.column_dimensions[get_column_letter(col_index["Run Mode"])].width       = 14
    ws.column_dimensions[get_column_letter(col_index["Backtest Start"])].width = 13
    ws.column_dimensions[get_column_letter(col_index["Backtest End"])].width   = 13
    ws.column_dimensions[get_column_letter(col_index["OOS Start"])].width      = 13
    ws.column_dimensions[get_column_letter(col_index["Pricing Model"])].width  = 13
    ws.column_dimensions[get_column_letter(col_index["Tx Cost %"])].width      = 10
    ws.column_dimensions[get_column_letter(col_index["Tax Drag %"])].width     = 10
    ws.column_dimensions[get_column_letter(col_index["Data Freq"])].width      = 10
    ws.column_dimensions[get_column_letter(col_index["Tickers"])].width        = 45
    ws.column_dimensions[get_column_letter(col_index["Results Folder"])].width = 55
    for strategy in STRATEGY_NAMES:
        for metric in METRIC_COLS:
            flat = f"{strategy}_{metric}"
            ws.column_dimensions[get_column_letter(col_index[flat])].width = 11

    # Freeze the two header rows and the meta columns
    freeze_col = get_column_letter(len(META_COLS) + 1)
    ws.freeze_panes = f"{freeze_col}3"

    wb.save(log_path)


def append_to_master_log(results_dir: str,
                         stats_list: list[StrategyStats],
                         weights: dict,
                         label: str):
    """
    Append one row to results/master_log.xlsx.
    Rewrites the full file each time to maintain correct formatting.
    On first run creates the file; on subsequent runs reads existing data,
    appends the new row, and rewrites.
    """
    log_path  = os.path.join("results", "master_log.xlsx")
    flat_cols = _all_flat_columns()
    new_row   = build_log_row(results_dir, stats_list, weights, label)

    os.makedirs("results", exist_ok=True)

    existing_rows = []
    if os.path.exists(log_path):
        # Read existing data rows (skip the two header rows).
        # Reindex to current flat_cols so legacy rows with fewer columns
        # get empty strings for any new columns rather than shifting data.
        existing_df = pd.read_excel(log_path, header=None, skiprows=2)
        existing_df.columns = flat_cols[:len(existing_df.columns)]
        existing_df = existing_df.reindex(columns=flat_cols, fill_value="")
        for _, r in existing_df.iterrows():
            existing_rows.append(
                {k: ("" if isinstance(v, float) and pd.isna(v) else v)
                 for k, v in r.to_dict().items()}
            )

    all_rows = existing_rows + [new_row]
    _write_excel_log(log_path, all_rows)

    print(f"  Master log updated -> {log_path}")


# ===========================================================================
# PRETTY PRINTING
# ===========================================================================

def print_header(title: str):
    """Print a clearly visible section divider to the terminal."""
    print("\n" + "=" * 60)
    print(f"  {title}")
    print("=" * 60)


def print_rebalancing(instructions: pd.DataFrame, total_value: float):
    """Print formatted rebalancing instructions to the terminal."""
    print_header("MONTHLY REBALANCING INSTRUCTIONS")
    print(f"  Total Portfolio Value: ${total_value:,.2f}\n")

    needs_action = instructions[instructions["Action"] != "HOLD"]
    hold_tickers = instructions[instructions["Action"] == "HOLD"]["Ticker"].tolist()

    if needs_action.empty:
        print("  No rebalancing needed -- all allocations within threshold.\n")
    else:
        for _, row in needs_action.iterrows():
            tag = "SELL" if row["Action"] == "SELL" else "BUY "
            print(f"  {tag}  {row['Ticker']:5s}  ${row['$ Amount']:>8,.2f}"
                  f"   (current {row['Current Weight']:.1f}%  ->"
                  f"  target {row['Target Weight']:.1f}%)")

    if hold_tickers:
        print(f"\n  HOLD: {', '.join(hold_tickers)}")

    if not needs_action.empty:
        print(f"\n  Full breakdown:")
        print(instructions[["Ticker", "Current Weight", "Target Weight",
                            "Drift (%)", "Action", "$ Amount"]].to_string(index=False))


def print_stats(stats_list: list[StrategyStats]):
    """Print formatted performance statistics to the terminal."""
    print_header("PERFORMANCE STATISTICS")
    for s in stats_list:
        print(f"\n  --- {s.name} ---")
        print(f"  {'Period (years)':<25} {s.period_years}")
        print(f"  {'CAGR (%)':<25} {s.cagr}")
        print(f"  {'Max Drawdown (%)':<25} {s.max_drawdown}")
        print(f"  {'Avg Drawdown (%)':<25} {s.avg_drawdown}")
        print(f"  {'Max DD Duration (months)':<25} {s.max_dd_duration}")
        print(f"  {'Avg Recovery (months)':<25} {s.avg_recovery_time}")
        print(f"  {'Ulcer Index':<25} {s.ulcer_index}")
        print(f"  {'Sharpe Ratio':<25} {s.sharpe}")
        print(f"  {'Sortino Ratio':<25} {s.sortino}")
        print(f"  {'Calmar Ratio':<25} {s.calmar}")
        print(f"  {'Final Value ($)':<25} {s.final_value}")