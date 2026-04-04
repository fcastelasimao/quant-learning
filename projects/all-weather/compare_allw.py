"""
compare_allw.py
===============
Head-to-head comparison of validated All Weather strategies vs the
Bridgewater ALLW ETF (launched March 2025, ticker ALLW on NASDAQ).

All metrics are computed from DAILY prices over the ALLW overlap window,
so MaxDD and Ulcer Index reflect intraday drawdowns, not monthly snapshots.

Fee-adjusted rows apply annualised expense-ratio drag to the gross series:
  ALLW: 0.85% p.a.   DIY strategies: 0.12% p.a. (weighted avg ETF expense)

Outputs
-------
  - Side-by-side performance table (stdout)
  - results/allw_comparison_growth.png
  - results/allw_comparison_*.xlsx

Usage
-----
  conda run -n allweather python3 compare_allw.py
"""

from __future__ import annotations

import json
import os
import sys
import warnings
from dataclasses import dataclass
from datetime import date, datetime

import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import matplotlib.dates as mdates
import numpy as np
import pandas as pd
import yfinance as yf
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore", category=FutureWarning)
warnings.filterwarnings("ignore", message=".*auto_adjust.*")

_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_RESULTS_DIR = os.path.join(_SCRIPT_DIR, "results")
os.makedirs(_RESULTS_DIR, exist_ok=True)

# ---------------------------------------------------------------------------
# CONSTANTS
# ---------------------------------------------------------------------------

DATE_START = "2025-03-06"  # ALLW launch date
#DATE_START = "2026-01-01"

#DATE_END   = "2025-06-30"
DATE_END   = date.today().strftime("%Y-%m-%d")

ALLW_FEE = 0.0085    # Bridgewater ALLW annual expense ratio
DIY_FEE  = 0.0012    # Weighted avg ETF expense ratio for DIY strategies

# 60/40 benchmark weights
SIXTY_FORTY_EQUITY = 0.60
SIXTY_FORTY_BOND   = 0.40

IRAN_WAR_DATE = pd.Timestamp("2026-02-28")
WATERMARK     = "github.com/fcastelasimao/quant-learning"

# DARK THEME COLOURS (charts only) ---------------------------------------------

DARK_BG     = "#0d1117"
PANEL_BG    = "#161b22"
GRID_COL    = "#30363d"
TEXT_COL    = "#c9d1d9"
BORDER_COL  = "#30363d"

# ---------------------------------------------------------------------------
# LOAD ALLOCATIONS FROM strategies.json
# ---------------------------------------------------------------------------

def _load_strategies_json() -> dict:
    """Load strategies.json from the same directory as this script."""
    base = os.path.dirname(os.path.abspath(__file__))
    path = os.path.join(base, "strategies.json")
    if not os.path.exists(path):
        print(f"WARNING: {path} not found — using hardcoded fallback allocations")
        return {}
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f).get("strategies", {})

_STRATEGIES_JSON = _load_strategies_json()


def _alloc_from_json(strategy_id: str, use_live: bool = False) -> dict[str, float]:
    """
    Read allocation from strategies.json, optionally translating to live tickers.
    Falls back to empty dict if strategy_id not found (caller should handle).
    """
    payload = _STRATEGIES_JSON.get(strategy_id)
    if payload is None:
        return {}
    alloc = dict(payload["allocation"])
    if use_live:
        live_map = payload.get("live_tickers", {})
        translated = {}
        for ticker, weight in alloc.items():
            tradable = live_map.get(ticker, ticker)
            translated[tradable] = translated.get(tradable, 0.0) + weight
        return translated
    return alloc

# ---------------------------------------------------------------------------
# STRATEGY & BENCHMARK REGISTRY
# ---------------------------------------------------------------------------
# Every strategy is defined ONCE here. Allocations are read from
# strategies.json. To add a new strategy, add one entry below and
# (if new) one entry in strategies.json. Set enabled=False to skip.

@dataclass
class StrategyDef:
    """Single source of truth for a portfolio strategy."""
    key: str                            # unique identifier
    label: str                          # display name for tables & legends
    strategy_id: str                    # key in strategies.json
    use_live_tickers: bool              # use live_tickers translation?
    fee: float                          # annual expense ratio
    color: str                          # hex colour for charts
    allocation: dict[str, float] = None # populated from strategies.json at startup
    enabled: bool = True                # False = skip entirely
    line_width: float = 2.2             # matplotlib line width
    line_style: str = "-"               # matplotlib line style
    rebalance: str = "monthly"          # "monthly" | "buy_and_hold"
    show_in_chart: bool = True          # include in growth chart?
    chart_order: int = 50               # lower = plotted first (back layer)


@dataclass
class BenchmarkDef:
    """Chart-only benchmark (no allocation dict, built ad-hoc in main)."""
    key: str
    label: str
    color: str
    line_width: float = 1.5
    line_style: str = "-"
    chart_order: int = 90


STRATEGY_REGISTRY: list[StrategyDef] = [
    # ── 6-asset manual (equal-ish weights) ────────────────────────────────
    StrategyDef(
        key="6asset_manual_live", label="6asset_manual_live",
        strategy_id="6asset_tip_gsg", use_live_tickers=True,
        fee=DIY_FEE, color="#58a6ff",
        enabled=False,
        line_width=2.2, line_style="--", show_in_chart=False, chart_order=10,
    ),
    StrategyDef(
        key="6asset_manual_backtest", label="6asset_manual_backtest",
        strategy_id="6asset_tip_gsg", use_live_tickers=False,
        fee=DIY_FEE, color="#FF1919",
        enabled=False,
        line_width=1.5, line_style="--", show_in_chart=False, chart_order=11,
    ),
    # ── 6-asset risk-parity average (production weights) ──────────────────
    StrategyDef(
        key="6asset_rpavg_live", label="6asset_rpavg_live",
        strategy_id="6asset_tip_gsg_rpavg", use_live_tickers=True,
        fee=DIY_FEE, color="#d2a8ff",
        line_width=2.2, line_style="-", show_in_chart=True, chart_order=30,
    ),
    StrategyDef(
        key="6asset_rpavg_backtest", label="My_strategy",
        strategy_id="6asset_tip_gsg_rpavg", use_live_tickers=False,
        fee=DIY_FEE, color="#58a6ff",
        line_width=2.2, line_style="-", show_in_chart=True, chart_order=31,
    ),
    StrategyDef(
        key="6asset_rpavg_bh", label="My_strategy BUY&HOLD",
        strategy_id="6asset_tip_gsg_rpavg", use_live_tickers=False,
        fee=DIY_FEE, color="#ff7b72",
        line_width=1.8, line_style="--",
        rebalance="buy_and_hold", show_in_chart=True, chart_order=32,
    ),
    # ── 8-asset (archived — 6-asset wins on all metrics) ───────────────────
    StrategyDef(
        key="8asset_backtest", label="8asset_backtest",
        strategy_id="8asset_CPER_DBA_GLD_IEF_IJR_QQQ_SPY_TLT", use_live_tickers=False,
        fee=DIY_FEE, color="#fc58ff",
        enabled=False,
        line_width=2.2, line_style="--", show_in_chart=False, chart_order=34,
    ),
    # ── 5-asset Dalio classic ─────────────────────────────────────────────
    StrategyDef(
        key="5asset_dalio_live", label="5asset_dalio_live",
        strategy_id="5asset_dalio", use_live_tickers=True,
        fee=DIY_FEE, color="#3ddc97",
        enabled=False,
        line_width=1.5, line_style="-.", show_in_chart=False, chart_order=40,
    ),
    StrategyDef(
        key="5asset_dalio_backtest", label="5asset_dalio (Dalio classic)",
        strategy_id="5asset_dalio", use_live_tickers=False,
        fee=DIY_FEE, color="#3ddc97",
        line_width=1.5, line_style="-.", show_in_chart=True, chart_order=41,
    ),

]

BENCHMARK_REGISTRY: list[BenchmarkDef] = [
    BenchmarkDef("ALLW",  "ALLW (Bridgewater, 0.85% fee)", "#f0b429", 2.2, "-",  80),
    BenchmarkDef("SPY",   "SPY (S&P 500)",                 "#f78166", 1.5, ":",  90),
    BenchmarkDef("60/40", "60/40 (SPY/TLT)",               "#3fb950", 1.5, "-.", 95),
]


def _resolve_allocations() -> None:
    """Populate each StrategyDef.allocation from strategies.json at import time."""
    for s in STRATEGY_REGISTRY:
        if not s.enabled:
            continue
        alloc = _alloc_from_json(s.strategy_id, use_live=s.use_live_tickers)
        if not alloc:
            print(f"WARNING: strategy_id '{s.strategy_id}' not found in strategies.json — disabling {s.key}")
            s.enabled = False
        s.allocation = alloc  # type: ignore[attr-defined]

_resolve_allocations()

# ── Derived ticker lists (auto-computed from enabled strategies) ──────────
ALL_TICKERS = sorted({t for s in STRATEGY_REGISTRY if s.enabled for t in s.allocation})
# SPY and TLT are always needed: SPY as benchmark, TLT for the 60/40 series.
FETCH_TICKERS = sorted(set(ALL_TICKERS + ["ALLW", "SPY", "TLT"]))


# DATA FETCHING ------------------------------------------------------------------

def _strip_tz(idx: pd.DatetimeIndex) -> pd.DatetimeIndex:
    """Remove timezone info so all indices compare cleanly."""
    if idx.tz is not None:
        return idx.tz_localize(None)
    return idx

def fetch_daily_prices(start: str = DATE_START,
                       end:   str = DATE_END) -> pd.DataFrame:
    """
    Fetch daily adjusted close prices for all comparison tickers.

    Returns a DataFrame indexed by timezone-naive date.
    Raises SystemExit with a clear message if ALLW data is unavailable.
    """
    print(f"Fetching daily prices | {' '.join(FETCH_TICKERS)}")
    print(f"  Period: {start}  →  {end}")

    frames = {}
    failed = []
    for ticker in FETCH_TICKERS:
        try:
            t    = yf.Ticker(ticker)
            hist = t.history(start=start, end=end, auto_adjust=True)
            if hist.empty:
                failed.append(ticker)
                continue
            hist.index = _strip_tz(hist.index)
            frames[ticker] = hist["Close"].rename(ticker)
        except Exception as exc:
            print(f"  ERROR fetching {ticker}: {exc}")
            failed.append(ticker)

    if "ALLW" in failed:
        print(
            "\n" + "=" * 60 + "\n"
            "DATA UNAVAILABLE: ALLW price history could not be retrieved\n"
            "from Yahoo Finance.  ALLW (Bridgewater All Weather ETF) was\n"
            "launched in March 2025.  If yfinance does not yet carry it,\n"
            "try again later or source the data manually from:\n"
            "  https://finance.yahoo.com/quote/ALLW/history\n"
            "=" * 60
        )
        sys.exit(1)

    if failed:
        print(f"=== > WARNING: could not fetch {failed}; proceeding without them.")

    prices = pd.DataFrame(frames).dropna(how="all")
    prices = prices.ffill()

    print(f"  {len(prices)} trading days loaded"
          f"  ({prices.index[0].date()} → {prices.index[-1].date()})\n")
    return prices


# DAILY PORTFOLIO SERIES -----------------------------------------------------------

def build_daily_series_buy_and_hold(prices: pd.DataFrame,
                                    allocation: dict,
                                    start_value: float = 100.0) -> pd.Series:
    """
    Compute daily portfolio value using a buy-and-hold strategy
    (initial weights, no rebalancing). Starting value = start_value.
    """
    tickers  = [t for t in allocation if t in prices.columns]
    alloc    = {t: allocation[t] for t in tickers}
    total_w  = sum(alloc.values())
    alloc    = {t: w / total_w for t, w in alloc.items()}

    first    = prices[tickers].iloc[0]
    shares   = {t: (start_value * w) / float(first[t]) for t, w in alloc.items()}
    daily    = sum(shares[t] * prices[t] for t in tickers)
    return daily.rename("portfolio")

def build_daily_series(prices: pd.DataFrame,
                       allocation: dict,
                       start_value: float = 100.0,
                       rebalance: bool = True) -> pd.Series:
    """
    Compute a daily portfolio value series from an allocation dict.

    When *rebalance* is True, weights are restored to targets at each
    month-end.  When False, the portfolio drifts (buy-and-hold).
    """
    tickers = [t for t in allocation if t in prices.columns]
    alloc = {t: allocation[t] for t in tickers}
    total_w = sum(alloc.values())
    alloc = {t: w / total_w for t, w in alloc.items()}

    first = prices[tickers].iloc[0]
    shares = {t: (start_value * w) / float(first[t]) for t, w in alloc.items()}

    if not rebalance:
        daily = sum(shares[t] * prices[t] for t in tickers)
        return daily.rename("portfolio")

    # Monthly rebalancing: restore target weights at each month-end
    month_ends = set(prices.resample("ME").last().dropna(how="all").index)
    values = []

    for date, row in prices[tickers].iterrows():
        port_val = sum(shares[t] * float(row[t]) for t in tickers)
        values.append(port_val)

        if date in month_ends:
            for t, w in alloc.items():
                shares[t] = (port_val * w) / float(row[t])

    return pd.Series(values, index=prices[tickers].index, name="portfolio")

def apply_annual_fee(series: pd.Series, annual_fee: float) -> pd.Series:
    """Apply a compounding annual expense-ratio drag to a value series."""
    n_days   = np.arange(len(series))
    discount = (1.0 - annual_fee) ** (n_days / 252.0)
    return series * discount


# DAILY STATISTICS -------------------------------------------------------------

def _daily_stats(series: pd.Series,
                 label:  str,
                 allocation: dict = None,
                 annualise: int = 252) -> dict:
    """Compute all comparison metrics from a daily value series."""
    s     = series.dropna()
    if len(s) < 2:
        return {k: float("nan") for k in
                ["label", "total_ret", "cagr", "max_dd", "calmar",
                 "ulcer", "sortino", "vol", "worst_month", "best_month",
                 "n_days"]}

    n_days   = len(s)
    years    = n_days / annualise
    total_r  = (s.iloc[-1] / s.iloc[0] - 1) * 100
    cagr     = ((s.iloc[-1] / s.iloc[0]) ** (1 / years) - 1) * 100

    peak     = s.cummax()
    dd_pct   = ((s - peak) / peak) * 100
    max_dd   = dd_pct.min()

    calmar   = cagr / abs(max_dd) if abs(max_dd) > 1e-6 else 0.0
    ulcer    = float(np.sqrt((dd_pct ** 2).mean()))

    rets     = s.pct_change().dropna()
    down     = rets[rets < 0]
    if len(down) == 0 or down.std() < 1e-10:
        sortino = 0.0
    else:
        sortino = float((rets.mean() / down.std()) * np.sqrt(annualise))

    vol = float(rets.std() * np.sqrt(annualise) * 100)

    monthly_prices = s.resample("ME").last()
    monthly_rets   = monthly_prices.pct_change().dropna() * 100
    worst_month    = float(monthly_rets.min()) if len(monthly_rets) > 0 else float("nan")
    best_month     = float(monthly_rets.max()) if len(monthly_rets) > 0 else float("nan")

    allocs = " | ".join(f"{t}={w:.1%}" for t, w in allocation.items()) if allocation is not None else ""

    return {
        "label":       label,
        "allocations": allocs,
        "total_ret":   round(total_r,    2),
        "cagr":        round(cagr,       2),
        "max_dd":      round(max_dd,     2),
        "calmar":      round(calmar,     3),
        "ulcer":       round(ulcer,      3),
        "sortino":     round(sortino,    3),
        "vol":         round(vol,        2),
        "worst_month": round(worst_month, 2),
        "best_month":  round(best_month,  2),
        "n_days":      n_days,
    }

# PRINT TABLE --------------------------------------------------------------------------

def print_comparison_table(rows: list[dict],
                            period_label: str) -> None:
    """Print a formatted side-by-side comparison table to stdout."""
    SEP  = "─" * 90
    HDR  = f"{'Strategy':<35} {"Allocations":<75} {'TotRet':>8} {'CAGR':>8} {'MaxDD':>8} "  \
           f"{'Calmar':>8} {'Ulcer':>8} {'Sortino':>9}"

    print()
    print(SEP)
    print(f"  ALLW HEAD-TO-HEAD COMPARISON  ─  {period_label}")
    print(SEP)
    print(HDR)
    print(SEP)

    for r in rows:
        if r is None:
            print()
            continue
        label    = r["label"]
        allocation = r["allocations"]
        tot_r    = f"{r['total_ret']:>+.2f}%"
        cagr     = f"{r['cagr']:>+.2f}%"
        max_dd   = f"{r['max_dd']:.2f}%"
        calmar   = f"{r['calmar']:.3f}"
        ulcer    = f"{r['ulcer']:.3f}"
        sortino  = f"{r['sortino']:+.3f}"
        print(f"  {label:<35}  {allocation:<75}  {tot_r:>8} {cagr:>8} {max_dd:>9} " \
              f"{calmar:>8} {ulcer:>8} {sortino:>9}")

    print(SEP)
    print(f"  Metrics computed from daily prices  |  MaxDD = daily peak-to-trough")
    print(f"  Fee-adjusted rows apply annualised expense-ratio drag to gross returns")
    print(f"    ALLW: 0.85% p.a.  |  DIY (6asset / 7asset): 0.12% p.a.")
    print(SEP)
    print()

# DARK THEME HELPER ---------------------------------------------------------------

def _style_ax(ax: plt.Axes) -> None:
    ax.set_facecolor(PANEL_BG)
    ax.tick_params(colors=TEXT_COL, labelsize=9)
    for sp in ax.spines.values():
        sp.set_color(BORDER_COL)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.yaxis.label.set_color(TEXT_COL)
    ax.xaxis.label.set_color(TEXT_COL)
    ax.title.set_color("white")
    ax.grid(axis="y", color=GRID_COL, alpha=0.45, linewidth=0.7)
    ax.grid(axis="x", color=GRID_COL, alpha=0.25, linewidth=0.4)


def _add_watermark(ax: plt.Axes) -> None:
    ax.text(
        0.995, 0.012, WATERMARK,
        transform=ax.transAxes,
        ha="right", va="bottom",
        fontsize=7, color="#444c56", alpha=0.8,
        style="italic",
    )

def _save_both(fig: plt.Figure, base_name: str) -> None:
    path = os.path.join(_RESULTS_DIR, f"{base_name}.png")
    plt.savefig(path, dpi=100, bbox_inches="tight",
                facecolor=fig.get_facecolor())
    print(f"  Saved → {path}")


# CHART 1: GROWTH CHART ----------------------------------------------------------------

def plot_growth_chart(daily_series: dict[str, pd.Series]) -> None:
    """
    Cumulative growth chart: $10,000 invested on ALLW launch date.
    """
    print("Building growth chart ...")

    allw_start = daily_series["ALLW"].first_valid_index()

    fig, ax = plt.subplots(figsize=(12, 6.75))
    fig.patch.set_facecolor(DARK_BG)
    _style_ax(ax)

    # Build plot order from registries
    plot_items: list[tuple[str, str, float, str, str, int]] = []
    for s in STRATEGY_REGISTRY:
        if s.enabled and s.show_in_chart:
            plot_items.append((s.key, s.color, s.line_width, s.line_style, s.label, s.chart_order))
    for b in BENCHMARK_REGISTRY:
        plot_items.append((b.key, b.color, b.line_width, b.line_style, b.label, b.chart_order))
    plot_items.sort(key=lambda x: x[5])

    for key, color, lw, ls, label, _ in plot_items:
        if key not in daily_series:
            continue
        raw = daily_series[key].dropna()
        s   = raw.loc[allw_start:]
        s   = s / s.iloc[0] * 10_000

        ax.plot(s.index, s.values, color=color, lw=lw, linestyle=ls,
                label=label, alpha=0.92)

        final_val  = s.iloc[-1]
        final_date = s.index[-1]
        ax.annotate(
            f"${final_val:,.0f}",
            xy=(final_date, final_val),
            xytext=(8, 0),
            textcoords="offset points",
            color=color,
            fontsize=9,
            fontweight="bold",
            va="center",
        )

    # Iran war annotation
    ref = daily_series.get("SPY", daily_series.get("ALLW"))
    if ref is not None:
        s_ref = ref.loc[allw_start:]
        if s_ref.index[0] <= IRAN_WAR_DATE <= s_ref.index[-1]:
            ax.axvline(IRAN_WAR_DATE, color="#ff7b72", lw=1.2,
                       linestyle="--", alpha=0.75)
            ymin, ymax = ax.get_ylim()
            ax.text(
                IRAN_WAR_DATE, ymin + (ymax - ymin) * 0.04,
                " Iran war\n starts",
                color="#ff7b72", fontsize=8.5, va="bottom",
                style="italic", alpha=0.85,
            )

    ax.set_title(
        "$10,000 invested in March 2025 — where are you today?",
        fontsize=12, pad=12, color="white",
    )
    ax.set_ylabel("Portfolio Value ($)", fontsize=10)
    ax.yaxis.set_major_formatter(
        mticker.FuncFormatter(lambda x, _: f"${x:,.0f}"))
    ax.xaxis.set_major_formatter(mdates.DateFormatter("%b %Y"))
    ax.xaxis.set_major_locator(mdates.MonthLocator(interval=2))
    plt.setp(ax.xaxis.get_majorticklabels(), rotation=25, ha="right")

    ax.legend(
        fontsize=9, facecolor="#21262d", edgecolor=BORDER_COL,
        labelcolor=TEXT_COL, loc="upper left", framealpha=0.9,
    )

    _add_watermark(ax)
    plt.tight_layout(pad=1.5)
    _save_both(fig, f"{date.today().strftime('%Y-%m-%d')}_allw_comparison_growth_from{DATE_START}_to{DATE_END}")
    plt.close(fig)


# CHART 2: CUMULATIVE FEE DRAG PROJECTION -----------------------------------------

def plot_fee_drag_chart() -> None:
    """Projected cost of fees on a $100k portfolio over 30 years."""
    print("Building fee drag chart ...")

    STARTING_VALUE = 100_000
    GROSS_RETURN   = 0.07
    YEARS          = np.linspace(0, 30, 1000)
    MARKS          = [10, 20, 30]

    diy_color = "#58a6ff"
    allw_color = "#f0b429"
    for s in STRATEGY_REGISTRY:
        if "rpavg" in s.key and "live" in s.key:
            diy_color = s.color
            break
    for b in BENCHMARK_REGISTRY:
        if b.key == "ALLW":
            allw_color = b.color
            break

    gross     = STARTING_VALUE * (1 + GROSS_RETURN) ** YEARS
    allw_net  = STARTING_VALUE * (1 + GROSS_RETURN - ALLW_FEE) ** YEARS
    diy_net   = STARTING_VALUE * (1 + GROSS_RETURN - DIY_FEE)  ** YEARS

    fig, ax = plt.subplots(figsize=(12, 6.75))
    fig.patch.set_facecolor(DARK_BG)
    _style_ax(ax)

    ax.plot(YEARS, gross,    color="#8b949e", lw=1.5, linestyle=":",
            label="Gross (no fees) — 7.00% p.a.", alpha=0.7)
    ax.plot(YEARS, diy_net,  color=diy_color, lw=2.5,
            label=f"DIY all-weather — 0.12% p.a. fee (net {GROSS_RETURN - DIY_FEE:.2%})")
    ax.plot(YEARS, allw_net, color=allw_color, lw=2.5,
            label=f"ALLW ETF        — 0.85% p.a. fee (net {GROSS_RETURN - ALLW_FEE:.2%})")

    ax.fill_between(YEARS, diy_net, allw_net,
                    color="#f0b429", alpha=0.08, label="Fee drag gap")

    for yr in MARKS:
        idx    = np.searchsorted(YEARS, yr)
        diy_v  = diy_net[idx]
        allw_v = allw_net[idx]
        gap    = diy_v - allw_v

        ax.plot(yr, diy_v,  "o", color=diy_color,  ms=7, zorder=5)
        ax.plot(yr, allw_v, "o", color=allw_color,  ms=7, zorder=5)

        ax.annotate(
            f"${diy_v/1e3:.0f}k",
            xy=(yr, diy_v), xytext=(yr + 0.4, diy_v * 1.035),
            color=diy_color, fontsize=9, fontweight="bold",
            arrowprops=dict(arrowstyle="-", color=diy_color, lw=0.7),
        )
        ax.annotate(
            f"${allw_v/1e3:.0f}k",
            xy=(yr, allw_v), xytext=(yr + 0.4, allw_v * 0.93),
            color=allw_color, fontsize=9, fontweight="bold",
            arrowprops=dict(arrowstyle="-", color=allw_color, lw=0.7),
        )
        ax.text(
            yr - 0.6, (diy_v + allw_v) / 2,
            f"−${gap/1e3:.0f}k\ngap",
            color="#8b949e", fontsize=7.5, ha="right", va="center",
            style="italic",
        )

    ax.set_title(
        "Cumulative Fee Drag — ALLW (0.85%) vs DIY All Weather (0.12%)\n"
        f"Starting value: ${STARTING_VALUE:,.0f}   Assumed gross return: {GROSS_RETURN:.0%} p.a.",
        fontsize=11, pad=12, color="white",
    )
    ax.set_xlabel("Years", fontsize=10)
    ax.set_ylabel("Portfolio Value ($)", fontsize=10)
    ax.set_xlim(0, 31)
    ax.yaxis.set_major_formatter(
        mticker.FuncFormatter(lambda x, _: f"${x/1e3:,.0f}k"))

    ax.legend(
        fontsize=9.5, facecolor="#21262d", edgecolor=BORDER_COL,
        labelcolor=TEXT_COL, loc="upper left", framealpha=0.92,
    )

    _add_watermark(ax)
    plt.tight_layout(pad=1.5)
    _save_both(fig, f"{date.today().strftime('%Y-%m-%d')}_allw_fee_drag_from{DATE_START}_to{DATE_END}")
    plt.close(fig)


# EXCEL EXPORT -----------------------------------------------------------------

_EXCEL_COLS = [
    ("Strategy",      "label",       30,  None),
    ("Allocations",   "allocations", 70,  None),
    ("Total Ret (%)", "total_ret",   13,  "0.00"),
    ("CAGR (%)",      "cagr",        11,  "0.00"),
    ("Max DD (%)",    "max_dd",      11,  "0.00"),
    ("Calmar",        "calmar",      10,  "0.000"),
    ("Ulcer",         "ulcer",       10,  "0.000"),
    ("Sortino",       "sortino",     10,  "+0.000;-0.000"),
    ("Vol (%)",       "vol",         10,  "0.00"),
    ("Worst Mo (%)",  "worst_month", 13,  "0.00"),
    ("Best Mo (%)",   "best_month",  13,  "0.00"),
]


_HEADER_FILL = PatternFill("solid", fgColor="D9D9D9")  # light grey


def _xfont(bold: bool = False, italic: bool = False, size: int = 10) -> Font:
    return Font(bold=bold, italic=italic, size=size)


def _xborder() -> Border:
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def save_comparison_excel(rows: list, period_label: str) -> None:
    """
    Write the comparison table to results/allw_comparison_<YYYYMMDD>.xlsx.
    Clean layout: no background colours, empty row between strategy groups.
    """
    today    = datetime.now()
    filename = f"{date.today().strftime('%Y-%m-%d')}_allw_comparison_from{DATE_START}_to{DATE_END}.xlsx"
    out_path = os.path.join(_RESULTS_DIR, filename)

    n_cols       = len(_EXCEL_COLS)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right_align  = Alignment(horizontal="right",  vertical="center")
    left_align   = Alignment(horizontal="left",   vertical="center")

    wb = Workbook()
    ws = wb.active
    ws.title = "ALLW Comparison"

    # ── Row 1: period label merged across all columns ────────────────────────
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1,   end_column=n_cols)
    cell           = ws.cell(row=1, column=1, value=period_label)
    cell.font      = _xfont(bold=True, size=11)
    cell.alignment = center_align
    cell.fill      = _HEADER_FILL

    # ── Row 2: column headers ────────────────────────────────────────────────
    for col_i, (header, _, width, _) in enumerate(_EXCEL_COLS, start=1):
        cell           = ws.cell(row=2, column=col_i, value=header)
        cell.font      = _xfont(bold=True)
        cell.alignment = center_align
        cell.border    = _xborder()
        cell.fill      = _HEADER_FILL
        ws.column_dimensions[get_column_letter(col_i)].width = width

    # ── Data rows ────────────────────────────────────────────────────────────
    excel_row = 3

    for row in rows:
        if row is None:
            # Empty spacer row — still grey-fill col 1
            ws.cell(row=excel_row, column=1).fill = _HEADER_FILL
            excel_row += 1
            continue

        label      = row.get("label", "")
        is_fee_row = label.startswith("  ")

        for col_i, (_, key, _, num_fmt) in enumerate(_EXCEL_COLS, start=1):
            value = row.get(key, float("nan"))
            if isinstance(value, float) and pd.isna(value):
                value = ""

            cell        = ws.cell(row=excel_row, column=col_i, value=value)
            cell.border = _xborder()

            if col_i == 1:
                cell.font      = _xfont(bold=not is_fee_row, italic=is_fee_row)
                cell.alignment = left_align
                cell.fill      = _HEADER_FILL
            else:
                cell.font      = _xfont()
                cell.alignment = right_align
                if num_fmt and isinstance(value, (int, float)):
                    cell.number_format = num_fmt

        excel_row += 1

    # ── Footer ───────────────────────────────────────────────────────────────
    excel_row += 1
    footer_lines = [
        f"Generated: {today.strftime('%Y-%m-%d %H:%M')}",
        f"Period: {period_label}",
        ("Note: All metrics computed from daily prices. ALLW data from "
         "Yahoo Finance. Fee-adjusted rows apply annual expense ratio drag "
         "compounded daily. Rf = 0 assumed for Sortino."),
    ]
    small_font = _xfont(italic=True, size=9)
    for line in footer_lines:
        ws.merge_cells(start_row=excel_row, start_column=1,
                       end_row=excel_row,   end_column=n_cols)
        cell      = ws.cell(row=excel_row, column=1, value=line)
        cell.font = small_font
        excel_row += 1

    ws.freeze_panes = "B3"

    wb.save(out_path)
    print(f"  Excel saved → {out_path}")


# MAIN ----------------------------------------------------------------------

def main() -> None:
    # ── 1. Fetch data ────────────────────────────────────────────────────────
    prices = fetch_daily_prices(start=DATE_START, end=DATE_END)

    allw_start = prices["ALLW"].first_valid_index()
    prices     = prices.loc[allw_start:]
    period_label = (f"{prices.index[0].date()} → {prices.index[-1].date()}  "
                    f"({len(prices)} trading days,  "
                    f"{len(prices)/252:.2f} yrs)  ")

    # ── 2. Build daily portfolio value series from registry ─────────────────
    series = {}
    rows = []

    for strat in STRATEGY_REGISTRY:
        if not strat.enabled:
            continue

        if strat.rebalance == "monthly":
            s = build_daily_series(prices, strat.allocation, rebalance=True)
        else:
            s = build_daily_series_buy_and_hold(prices, strat.allocation)

        s_fa = apply_annual_fee(s, strat.fee)
        series[strat.key] = s
        gross_label = strat.label if "(gross)" in strat.label else f"{strat.label} (gross)"
        rows.append(_daily_stats(s, gross_label, allocation=strat.allocation))
        rows.append(_daily_stats(s_fa, f"  → fee-adj {strat.fee:.2%} p.a."))
        rows.append(None)  # spacer after every strategy+fee pair

    # ── Benchmarks ──────────────────────────────────────────────────────────
    s_allw   = prices["ALLW"] / prices["ALLW"].iloc[0] * 100.0
    s_spy    = prices["SPY"]  / prices["SPY"].iloc[0]  * 100.0
    if "TLT" in prices.columns:
        spy_sh   = SIXTY_FORTY_EQUITY / float(prices["SPY"].iloc[0])
        tlt_sh   = SIXTY_FORTY_BOND   / float(prices["TLT"].iloc[0])
        s_6040   = (spy_sh * prices["SPY"] + tlt_sh * prices["TLT"]) * 100.0
        s_6040   = s_6040 / s_6040.iloc[0] * 100.0
    else:
        s_6040 = None

    s_allw_fa = apply_annual_fee(s_allw, ALLW_FEE)

    rows.append(_daily_stats(s_allw,    "ALLW  (Bridgewater, gross)"))
    rows.append(_daily_stats(s_allw_fa, "  → fee-adj 0.85% p.a."))
    rows.append(None)
    rows.append(_daily_stats(s_spy,     "SPY  (benchmark)"))
    if s_6040 is not None:
        rows.append(_daily_stats(s_6040, "60/40  (SPY 60% / TLT 40%)"))

    # ── 3. Print table ───────────────────────────────────────────────────────
    print_comparison_table(rows, period_label)

    # ── 3b. Excel export ─────────────────────────────────────────────────────
    save_comparison_excel(rows, period_label)

    # ── 4. Chart ─────────────────────────────────────────────────────────────
    daily_series = {asset: s for asset, s in series.items()}
    daily_series["ALLW"] = s_allw
    daily_series["SPY"]  = s_spy

    if s_6040 is not None:
        daily_series["60/40"] = s_6040

    plot_growth_chart(daily_series)

    #plot_fee_drag_chart()

    print("\nDone.")


if __name__ == "__main__":
    main()
